
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages

from django.views.decorators.http import require_POST,require_GET
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.http import JsonResponse,HttpResponseRedirect,HttpResponse, HttpResponseBadRequest,FileResponse, Http404
from django.conf import settings
from django.template.loader import render_to_string,get_template
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone
from django.utils.encoding import escape_uri_path
from django.urls import reverse
from django.db.models import Q
from .forms import *

import xlrd
import math
from datetime import datetime,date,timedelta
from icecream import ic
from collections import defaultdict, deque, Counter
import pandas as pd
pd.set_option('display.max_rows', None)

import os, io, logging
import re
import json
from io import StringIO,BytesIO
import openpyxl
import xlwt
from openpyxl import Workbook
from pathlib import Path
from math import ceil
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration

# Create your views here.
def login_view(request):
    # 支持 next 参数跳转回原页面
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get("next") or request.POST.get("next") or reverse("project_config")
            return redirect(next_url)
        else:
            messages.error(request, "用户名或密码错误")
            # fallthrough -> re-render form
    return render(request, "dashboard/login.html")


def logout_view(request):
    logout(request)
    return redirect("dashboard_home")


@user_passes_test(lambda u: u.is_superuser)
def create_user(request):
    # 仅超级管理员可访问
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            new_user = form.save()
            messages.success(request, f"用户 {new_user.username} 创建成功")
            return redirect("project_config")
    else:
        form = UserCreationForm()
    return render(request, "dashboard/create_user.html", {"form": form})


def home(request):
    return render(request, "dashboard/index.html")

def user_manual(request):
    # 这里也可以做权限控制（如仅登录可见）
    return render(request, "dashboard/user_manual.html")

# 1 前端
# 前端功能入口
def frontend_entry(request):
    return render(request, 'dashboard/frontend/index.html')

# 关联后台参数配置中已设置的所有项目（获取唯一项目名及对应的第一条记录）
def get_project_list(request):
    # 获取唯一项目名及对应的第一条记录
    unique_projects = (
        SamplingConfiguration.objects
        .values('project_name')                # 只取项目名
        .distinct()                            # 去重（仅保留唯一项目名）
    )

    # 取每个唯一项目名的第一条记录（id）
    data = []
    for proj in unique_projects:
        first_obj = SamplingConfiguration.objects.filter(project_name=proj['project_name']).first()
        if first_obj:
            data.append({'id': first_obj.id, 'name': first_obj.project_name})

    return JsonResponse({'projects': data})

# 选择具体某个项目时，关联后台参数配置中该项目的配置信息
def get_project_detail(request, pk):
    configs = SamplingConfiguration.objects.get(pk=pk)
    data = {
        'id': configs.id,
        'name': configs.project_name,
        'default_upload_instrument': configs.default_upload_instrument
    }
    return JsonResponse(data)

@require_GET
def get_injection_plates(request):
    project_name  = request.GET.get("project_name", "").strip()
    instrument_num = request.GET.get("instrument_num", "").strip()
    systerm_num    = request.GET.get("systerm_num", "").strip()  # 新增

    plates = []
    if project_name and instrument_num:
        qs = InjectionPlateConfiguration.objects.filter(
            project_name=project_name,
            instrument_num=instrument_num,
        )
        if systerm_num:  # 传了就按系统号进一步过滤
            qs = qs.filter(systerm_num=systerm_num)

        # 兼容历史：可能有多条，合并去重
        for cfg in qs:
            raw = cfg.injection_plate
            if isinstance(raw, str):
                parts = [s.strip() for s in raw.split(",") if s.strip()]
            elif isinstance(raw, (list, tuple)):
                parts = list(raw)
            else:
                parts = []
            plates.extend(parts)

        # 去重并按字符串排序，避免顺序抖动
        plates = sorted(dict.fromkeys(plates), key=lambda x: str(x))

    return JsonResponse({"plates": plates})


@require_GET
def get_systerm_nums(request):
    project_name   = request.GET.get("project_name", "").strip()
    instrument_num = request.GET.get("instrument_num", "").strip()

    if not project_name or not instrument_num:
        return JsonResponse({"systerm_nums": []})

    # 取该项目 + 上机仪器下，已配置过的所有系统号（去重、排序）
    nums = (SamplingConfiguration.objects
            .filter(project_name=project_name, default_upload_instrument=instrument_num)
            .values_list("systerm_num", flat=True)
            .distinct())

    nums = sorted([n for n in nums if n])  # 过滤空值
    return JsonResponse({"systerm_nums": nums})


# Manual
def Manual_sampling(request):
    return render(request, 'dashboard/sampling/Manual.html')

# NIMBUS
def NIMBUS_sampling(request):
    return render(request, 'dashboard/sampling/NIMBUS.html')

# Starlet
def Starlet_sampling(request):
    return render(request, 'dashboard/sampling/Starlet.html')

def Starlet_qyzl(request):
    if request.method == 'POST' and request.FILES.get('input_file'):
        input_file = request.FILES['input_file']
        path = default_storage.save('tmp/' + input_file.name, ContentFile(input_file.read()))
        tmp_file = os.path.join(default_storage.location, path)

        wb = openpyxl.load_workbook(tmp_file, data_only=True)
        ws = wb.active

        AreaStartNum = int(ws['I1'].value)
        AreaRowNum = ws['K1'].value
        AreaStartPositionID = "B1"

        center_style = xlwt.XFStyle()
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        alignment.vert = xlwt.Alignment.VERT_CENTER
        center_style.alignment = alignment

        # 获取条码所在区域
        # Step 1: 自动检测“最大行和最大列”
        min_row, min_col = 3, 2  # B3
        max_row = ws.max_row
        max_col = ws.max_column

        # 1.1 找到从B3开始实际有数据的最大行和最大列
        actual_max_row, actual_max_col = min_row, min_col

        for row in range(min_row, ws.max_row+1):
            for col in range(min_col, ws.max_column+1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:  # 有值
                    if row > actual_max_row:
                        actual_max_row = row
                    if col > actual_max_col:
                        actual_max_col = col

        # Step 2: 从B3到实际最大行/列遍历，采集所有有值的条码，仍然按“按列优先”
        barcode_cells = ws.iter_cols(
            min_col=min_col, max_col=actual_max_col,
            min_row=min_row, max_row=actual_max_row
        )
        all_barcodes = [cell.value for col in barcode_cells for cell in col if cell.value]

        barcodes = all_barcodes[12:]
        OutputRowNum = len(barcodes)
        AreaNum = ceil(OutputRowNum / AreaRowNum)

        output = xlwt.Workbook()

        if AreaStartNum<=5:
            sheet = output.add_sheet("Import_1_5_Worklist")
        elif AreaStartNum<=10:
            sheet = output.add_sheet("Import_6_10_Worklist")
        elif AreaStartNum<=15:
            sheet = output.add_sheet("Import_11_15_Worklist")
        else:
            sheet = output.add_sheet("Import_16_20_Worklist")

        # sheet = output.add_sheet("Import_1_5_Worklist")

        sheet.col(1).width = 8 * 256
        sheet.col(2).width = 30 * 256
        sheet.col(3).width = 15 * 256
        sheet.col(4).width = 30 * 256
        sheet.col(5).width = 35 * 256
        sheet.col(6).width = 15 * 256
        sheet.col(7).width = 8 * 256

        headers = ["", "Index", "SourceLabwareID", "SourcePositionID", "SourceBarcode", "TargetLabwareID", "TargetPositionID", "Volume"]
        for col, head in enumerate(headers):
            sheet.write(0, col, head, center_style)

        def next_source_labware(n):
            return f"SMP_CAR_32_12x75_A00_{str(n).zfill(4)}"

        def get_target_labware_id(area_index, AreaStartNum):
            if AreaStartNum % 5 == 0:
                return f"Cos_96_DW_2mL_{str(5 + area_index).zfill(4)}"
            else:
                return f"Cos_96_DW_2mL_{str(AreaStartNum % 5 + area_index).zfill(4)}"

        def generate_target_position_ids(start, count, skip_list=None):
            if skip_list is None:
                skip_list = set()
            col_letter = start[0].upper()
            row_number = int(start[1:])
            result = []
            current_letter = col_letter
            current_number = row_number
            generated = 0
            while generated < count:
                pos = f"{current_letter}{current_number}"
                if pos not in skip_list:
                    result.append(pos)
                    generated += 1
                current_number += 1
                if current_number > 12:
                    current_number = 1
                    current_letter = chr(ord(current_letter) + 1)
            return result

        # -----------修正部分计数器--------------
        row_num = 1
        fixed_barcodes = [ws.cell(row=r, column=2).value for r in range(3, 15) if ws.cell(row=r, column=2).value]

        # 以下变量专用于“实际样本”（不含前12条曲线/质控）
        sample_labware_id_block = 1    # 只针对实际样本自增
        sample_position_id = 13        # 实际样本 SourcePositionID 从13开始

        for area_index in range(AreaNum):
            start_index = area_index * AreaRowNum
            end_index = min(start_index + AreaRowNum, OutputRowNum)
            area_barcodes = barcodes[start_index:end_index]

            total_barcodes = fixed_barcodes + area_barcodes

            start_row = row_num
            end_row = row_num + len(total_barcodes) - 1
            sheet.write_merge(start_row, end_row, 0, 0, f"{area_index + AreaStartNum}", center_style)
            target_labware_id = get_target_labware_id(area_index, AreaStartNum)

            fixed_positions = [f"A{n}" for n in range(1, 13)]
            skip_list = {f"B{AreaStartNum+area_index}"}
            dynamic_positions = generate_target_position_ids(AreaStartPositionID, len(area_barcodes), skip_list)
            total_positions = fixed_positions + dynamic_positions

            for index, (barcode, position) in enumerate(zip(total_barcodes, total_positions)):
                sheet.write(row_num, 1, index + 1, center_style)

                if index < 12:
                    # 前12条曲线/质控
                    sheet.write(row_num, 2, "SMP_CAR_32_12x75_A00_0001", center_style)
                    sheet.write(row_num, 3, index + 1, center_style)
                else:
                    # 实际样本部分
                    sheet.write(row_num, 2, next_source_labware(sample_labware_id_block), center_style)
                    sheet.write(row_num, 3, sample_position_id, center_style)

                    # 只在实际样本递增和判断
                    sample_position_id += 1
                    if sample_position_id > 32:
                        sample_position_id = 1
                        sample_labware_id_block += 1

                sheet.write(row_num, 4, barcode, center_style)
                sheet.write(row_num, 5, target_labware_id, center_style)
                sheet.write(row_num, 6, position, center_style)
                sheet.write(row_num, 7, 100, center_style)

                row_num += 1

        output_stream = BytesIO()
        output.save(output_stream)
        output_stream.seek(0)
        today_str = datetime.today().strftime("%Y-%m-%d")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"qyzl_plate_{AreaStartNum}_{AreaStartNum+AreaNum-1}_{timestamp}.xls"

        # 构造保存路径：Starlet/取样指令/YYYY-MM-DD/
        save_dir = os.path.join(settings.DOWNLOAD_ROOT, "Starlet", "取样指令", today_str)
        os.makedirs(save_dir, exist_ok=True)

        # 写入文件
        save_path = os.path.join(save_dir, file_name)
        with open(save_path, "wb") as f:
            f.write(output_stream.getvalue())

        # 生成“文件下载”页 URL（若你的 urls.py 没给这个路由起名，可直接用硬编码 '/dashboard/file_download/'）
        try:
            download_page_url = reverse("file_download")
        except Exception:
            download_page_url = "/dashboard/file_download/"

        # 直接文件 URL（静态暴露目录 + 平台/类别/日期/文件名）
        download_file_url = f"{settings.DOWNLOAD_URL}Starlet/取样指令/{today_str}/{file_name}"

        # 弹窗要显示的提示语
        popup_message = f"取样指令已生成：{file_name}（已保存至：Starlet/取样指令/{today_str}/）"

        return render(request, "dashboard/sampling/Starlet_qyzl.html", {
            "popup_message": popup_message,
            "download_page_url": download_page_url,
            "download_file_url": download_file_url,
        })

    return render(request, 'dashboard/sampling/Starlet_qyzl.html')


def Starlet_worksheet(request):
    return render(request, 'dashboard/sampling/Starlet_worksheet.html')

# Tecan
def Tecan_sampling(request):
    return render(request, 'dashboard/sampling/Tecan.html')

# 全血工作站
def WholeBloodWorkstation_sampling(request):
    return render(request, 'dashboard/sampling/WholeBloodWorkstation.html')

# 2 标本查找
def sample_search(request):
    return render(request, 'dashboard/sample_search/index.html')


def sample_search_stats_today(request):
    """
    返回：当天（record_date=today）的所有 project_name 的统计结果。
    统计口径和你给的 statistics(process_data) 基本一致：
      - 每个项目：根据 sample_name 自动提取字母前缀，或使用项目定制前缀集
      - 统计：实验号总数、起始/末尾实验号、空号（含区间描述）、共血、多血
    """
    today = datetime.now().date()
    qs = SampleRecord.objects.filter(record_date=today)

    # 若当天无数据，直接返回空结果
    if not qs.exists():
        return JsonResponse({"today_date": today.strftime("%Y-%m-%d"), "projects": []})

    # 可选：按项目设定“允许前缀”集合；项目未配置时自动从数据中抽取
    allowed_prefix_map = {
        "VAE": {"AE", "VF", "V", "ST"},
        "VD":  {"VD", "VF", "V", "YVD", "ST"},
        # 其他项目留空 → 自动提取
    }

    def process_project(records, allowed_prefixes=None):
        """
        records: 当个项目的 Queryset
        allowed_prefixes: set[str] | None
        返回：与示例一致的列表
        """
        # 收集候选 sample_name（兼容 'AE1234-VD5678' 这种 "-" 连接的情况）
        sample_names = []
        for r in records:
            if r.sample_name:
                parts = str(r.sample_name).split('-')
                if parts:
                    sample_names.append(parts[0])
                    if len(parts) > 1:
                        sample_names.append(parts[1])

        # 自动提取所有(字母+数字)的字母前缀，如 "VD123" → "VD"
        detected_prefixes = set()
        for name in sample_names:
            m = re.match(r"^[A-Za-z]+(?=\d)", str(name))
            if m:
                detected_prefixes.add(m.group())

        # 选择要使用的前缀集合
        prefixes = allowed_prefixes if allowed_prefixes else detected_prefixes

        # ★ 新增：统计时剔除 QC / STD 前缀
        prefixes = {p for p in prefixes if p not in {"QC", "STD"}}

        # 构造“条码→样本名集合”、“样本名→出现次数”，用于“共血/多血”
        barcode_to_samples = defaultdict(set)
        sample_count = defaultdict(int)
        for r in records:
            name_head = str(r.sample_name).split('-')[0] if r.sample_name else ""
            barcode_to_samples[r.barcode].add(name_head)
            sample_count[name_head] += 1

        # 生成统计行
        result_rows = []
        for prefix in sorted(prefixes):
            # 仅统计纯 '前缀+数字' 的样本名
            matching = [n for n in sample_names if re.fullmatch(fr"{prefix}\d+", str(n))]

            nums = sorted(
                int(re.sub(r"^[A-Za-z]+", "", n))
                for n in matching
                if re.sub(r"^[A-Za-z]+", "", n).isdigit()
            )

            total = len(matching)
            start_number = f"{prefix}{nums[0]}" if nums else None
            end_number   = f"{prefix}{nums[-1]}" if nums else None

            # 空号（找缺口）
            missing_ranges = []
            empty_count = 0
            for i in range(len(nums) - 1):
                gap = nums[i+1] - nums[i] - 1
                if gap > 0:
                    empty_count += gap
                    if gap == 1:
                        missing_ranges.append(f"{prefix}{nums[i] + 1}")
                    else:
                        missing_ranges.append(f"{prefix}{nums[i]+1}-{prefix}{nums[i+1]-1}")
            empty_info = ", ".join(missing_ranges)

            # 共血：同一个样本名映射多个条码（以样本名匹配、条码计数>1）
            shared_blood_samples = [
                name for name, barcodes in barcode_to_samples.items()
                if len(barcodes) > 1 and re.fullmatch(fr"{prefix}\d+", str(name))
            ]
            shared_blood_info  = ", ".join(sorted(set(shared_blood_samples)))
            shared_blood_count = len(set(shared_blood_samples))

            # 多血：同一个样本名在当天记录出现多次（>1）
            multi_blood_samples = [
                name for name, cnt in sample_count.items()
                if cnt > 1 and re.fullmatch(fr"{prefix}\d+", str(name))
            ]
            multi_blood_info  = ", ".join(sorted(set(multi_blood_samples)))
            multi_blood_count = len(set(multi_blood_samples))

            result_rows.append({
                "prefix":      prefix,
                "total":       total,
                "start":       start_number,
                "end":         end_number,
                "empty":       f"{empty_count}（{empty_info}）" if empty_count else "0",
                "sharedBlood": f"{shared_blood_count}（{shared_blood_info}）" if shared_blood_count else "0",
                "multiBlood":  f"{multi_blood_count}（{multi_blood_info}）" if multi_blood_count else "0",
            })
        return result_rows

    # 分项目统计
    projects_payload = []
    for proj in qs.values_list("project_name", flat=True).distinct().order_by("project_name"):
        proj_qs = qs.filter(project_name=proj)
        prefixes = allowed_prefix_map.get(proj)  # 未配置则为 None → 自动检测
        stats = process_project(proj_qs, prefixes)
        projects_payload.append({
            "project_name": proj,
            "statistics": stats
        })

    return JsonResponse({
        "today_date": today.strftime("%Y-%m-%d"),
        "projects": projects_payload
    })

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

def download_export(request, platform, date_name, project, filename):
    """
    统一的下载入口：无论 .pdf/.xlsx/.txt，都以 attachment 方式下载
    路径受限在 settings.DOWNLOAD_ROOT 下，防止越权访问
    """
    root = settings.DOWNLOAD_ROOT
    # 组装并规范化路径
    fpath = os.path.normpath(os.path.join(root, platform, date_name, project, filename))

    # 安全校验：必须在 DOWNLOAD_ROOT 内
    if not fpath.startswith(os.path.abspath(root) + os.sep):
        raise Http404("Invalid path")

    if not os.path.exists(fpath) or not os.path.isfile(fpath):
        raise Http404("File not found")

    # 统一用 FileResponse + as_attachment=True 强制下载
    # content_type 用二进制更稳妥，浏览器不会尝试内联展示
    resp = FileResponse(open(fpath, "rb"), as_attachment=True, filename=filename, content_type="application/octet-stream")

    # 兼容一些旧浏览器/中文文件名
    resp["Content-Disposition"] = f"attachment; filename*=UTF-8''{escape_uri_path(filename)}"
    return resp


def _should_hide_in_download_page(fname: str) -> bool:
    """
    文件下载页面不展示的文件类型
    - payload.json（用于重生成工作清单的数据源）
    """
    lower = (fname or "").lower().strip()
    # 兼容两种：固定名 payload.json / 以及你们这种 *.payload.json
    if lower == "payload.json" or lower.endswith(".payload.json"):
        return True
    return False


# 文件下载页面排序函数
TS_RE = re.compile(r"_(\d{8}_\d{6})_")  # 匹配 _20251219_094634_
def file_sort_key(fname: str):
    # 1) 提取时间戳：没有时间戳的放最后
    m = TS_RE.search(fname)
    ts = m.group(1) if m else "99999999_999999"

    # 2) 同一时间戳内的类型优先级：OnboardingList → WorkSheet → 其他
    if "OnboardingList" in fname:
        kind = 0
    elif "WorkSheet" in fname:
        kind = 1
    else:
        kind = 9

    # 3) 最后用文件名兜底，保证稳定排序
    return (ts, kind, fname)

HISTORY_DIRNAME = "历史文件"
STATION_DIRNAME = "岗位清单"

# 从 OnboardingList 文件名中提取项目名，例如：
# X6_OnboardingList_FXS-YZ04_S2_25OHD_20260108_060259_X6_GZ.txt -> 25OHD
_PROJECT_RE = re.compile(r"_S\d+_([A-Za-z0-9]+)_(\d{8})_")

def _extract_project_from_onboarding_filename(fname: str) -> str:
    s = (fname or "").strip()
    m = _PROJECT_RE.search(s)
    return (m.group(1) or "").strip() if m else ""

def _load_station_list_for_today() -> dict:
    """
    读取：DOWNLOAD_ROOT/岗位清单/{YYYY-MM-DD}/station_list.json
    返回：station_list["主条码->实验号列表"]（若不存在则返回空 dict）
    """
    today_str = timezone.localdate().strftime("%Y-%m-%d")
    fpath = os.path.join(settings.DOWNLOAD_ROOT, STATION_DIRNAME, today_str, "station_list.json")

    # ===== 临时排查用：打印实际查找路径与是否存在 =====
    print("station_list lookup path =", fpath)
    print("exists =", os.path.exists(fpath))
    # ==========================================
    
    if not os.path.exists(fpath):
        return {}

    try:
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}

    mapping = data.get("主条码->实验号列表") or {}
    return mapping if isinstance(mapping, dict) else {}


def _map_to_station_experiment(value: str, station_map: dict) -> tuple[bool, str, bool]:
    """
    无论 value 是什么，都尝试在 station_map 查实验号：
    - 查到：mapped = 第一个实验号，found=True
    - 查不到：mapped = 原输入 value，found=False
    规则：永远不阻断（ok 永远 True）
    返回：(ok, mapped_value, found)
    """
    v = (value or "").strip()
    if not v:
        return True, "", False

    exp_list = station_map.get(v)
    if isinstance(exp_list, list) and exp_list:
        mapped = str(exp_list[0]).strip()
        return True, mapped, True

    # 查不到：用原输入替换（不阻断）
    return True, v, False






def file_download(request):
    """
    展示 downloads 目录结构。
    - 通用：平台 / 日期 / 项目 / 文件
    - Starlet：平台 / {工作清单和上机列表 | 取样指令} / 日期 / (项目?) / 文件
    """
    root = settings.DOWNLOAD_ROOT
    os.makedirs(root, exist_ok=True)

    groups = []  # 输出给模板

    for platform in sorted(os.listdir(root)):  # 平台层
        # ✅ 新增：跳过历史文件目录和岗位清单目录（避免在主文件下载页显示）
        if platform == HISTORY_DIRNAME or platform == STATION_DIRNAME: 
            continue

        p_path = os.path.join(root, platform)
        if not os.path.isdir(p_path):
            continue

        # —— Starlet：多一层“类别”（工作清单和上机列表 / 取样指令）——
        if platform == "Starlet":
            category_names = ["工作清单和上机列表", "取样指令"]
            categories = []

            for cat in category_names:
                c_path = os.path.join(p_path, cat)
                if not os.path.isdir(c_path):
                    continue

                days = []
                # 日期倒序
                for date_name in sorted(os.listdir(c_path), reverse=True):
                    d_path = os.path.join(c_path, date_name)
                    if not (DATE_RE.match(date_name) and os.path.isdir(d_path)):
                        continue

                    # 判断“日期目录”下面是否还有项目层
                    proj_dirs = sorted([
                        s for s in os.listdir(d_path)
                        if os.path.isdir(os.path.join(d_path, s))
                    ])

                    if proj_dirs:
                        # 有项目层：平台 / 类别 / 日期 / 项目 / 文件
                        projects = []
                        for proj in proj_dirs:
                            proj_path = os.path.join(d_path, proj)
                            files = []
                            for fname in sorted(os.listdir(proj_path), key=file_sort_key):
                                if _should_hide_in_download_page(fname):    # ✅ 新增：隐藏 payload.json
                                    continue

                                fpath = os.path.join(proj_path, fname)
                                ext = fname.lower()
                                if not os.path.isfile(fpath):
                                    continue

                                files.append({
                                    "name": fname,
                                    "url": f"{settings.DOWNLOAD_URL}{platform}/{cat}/{date_name}/{proj}/{fname}",
                                    "force_download": ext.endswith((".pdf", ".txt", ".xlsx", ".xls", ".csv")),
                                })
                            projects.append({"name": proj, "files": files})

                        days.append({"date": date_name, "projects": projects})

                    else:
                        # 无项目层：平台 / 类别 / 日期 / 文件（用于“取样指令”）
                        files = []
                        for fname in sorted(os.listdir(d_path)):
                            if _should_hide_in_download_page(fname):    # ✅ 新增
                                continue

                            fpath = os.path.join(d_path, fname)
                            ext = fname.lower() 

                            if not os.path.isfile(fpath):
                                continue
                            files.append({
                                "name": fname,
                                "url": f"{settings.DOWNLOAD_URL}{platform}/{cat}/{date_name}/{fname}",
                                "force_download": ext.endswith((".pdf", ".txt", ".xlsx", ".xls", ".csv")),
                            })
                        days.append({"date": date_name, "files": files})

                categories.append({"category": cat, "days": days})

            groups.append({"group": platform, "categories": categories})
            continue  # Starlet 处理完，跳到下一个平台

        # —— 其他平台（沿用旧三层）：平台 / 日期 / 项目 / 文件 ——
        days = []
        for date_name in sorted(os.listdir(p_path), reverse=True):
            d_path = os.path.join(p_path, date_name)
            if not (DATE_RE.match(date_name) and os.path.isdir(d_path)):
                continue

            projects = []
            for proj in sorted(os.listdir(d_path)):
                proj_path = os.path.join(d_path, proj)
                if not os.path.isdir(proj_path):
                    continue

                files = []
                for fname in sorted(os.listdir(proj_path), key=file_sort_key):
                    if _should_hide_in_download_page(fname):    # ✅ 新增
                        continue

                    fpath = os.path.join(proj_path, fname)
                    ext = fname.lower()
                    if not os.path.isfile(fpath):
                        continue
                    files.append({
                        "name": fname,
                        "url": f"{settings.DOWNLOAD_URL}{platform}/{date_name}/{proj}/{fname}",
                        "force_download": ext.endswith((".pdf", ".txt", ".xlsx", ".xls", ".csv")),
                    })

                projects.append({"name": proj, "files": files})

            days.append({"date": date_name, "projects": projects})

        groups.append({"group": platform, "days": days})

    return render(request, "dashboard/file_download.html", {"groups": groups})


def file_download_history(request):
    """
    展示历史文件目录结构：DOWNLOAD_ROOT/历史文件/...
    结构与 file_download() 保持一致，便于复用模板渲染。
    """
    root = settings.DOWNLOAD_ROOT
    os.makedirs(root, exist_ok=True)

    hist_root = os.path.join(root, HISTORY_DIRNAME)
    os.makedirs(hist_root, exist_ok=True)

    groups = []

    for platform in sorted(os.listdir(hist_root)):  # 历史平台层
        p_path = os.path.join(hist_root, platform)
        if not os.path.isdir(p_path):
            continue

        # —— Starlet：多一层“类别”（工作清单和上机列表 / 取样指令）——
        if platform == "Starlet":
            category_names = ["工作清单和上机列表", "取样指令"]
            categories = []

            for cat in category_names:
                c_path = os.path.join(p_path, cat)
                if not os.path.isdir(c_path):
                    continue

                days = []
                for date_name in sorted(os.listdir(c_path), reverse=True):
                    d_path = os.path.join(c_path, date_name)
                    if not (DATE_RE.match(date_name) and os.path.isdir(d_path)):
                        continue

                    proj_dirs = sorted([
                        s for s in os.listdir(d_path)
                        if os.path.isdir(os.path.join(d_path, s))
                    ])

                    if proj_dirs:
                        projects = []
                        for proj in proj_dirs:
                            proj_path = os.path.join(d_path, proj)
                            files = []
                            for fname in sorted(os.listdir(proj_path), key=file_sort_key):
                                fpath = os.path.join(proj_path, fname)
                                ext = fname.lower()
                                if not os.path.isfile(fpath):
                                    continue
                                files.append({
                                    "name": fname,
                                    "url": f"{settings.DOWNLOAD_URL}{HISTORY_DIRNAME}/{platform}/{cat}/{date_name}/{proj}/{fname}",
                                    "force_download": ext.endswith((".pdf", ".txt", ".xlsx", ".xls", ".csv")),
                                })
                            projects.append({"name": proj, "files": files})
                        days.append({"date": date_name, "projects": projects})
                    else:
                        files = []
                        for fname in sorted(os.listdir(d_path)):
                            fpath = os.path.join(d_path, fname)
                            ext = fname.lower()
                            if not os.path.isfile(fpath):
                                continue
                            files.append({
                                "name": fname,
                                "url": f"{settings.DOWNLOAD_URL}{HISTORY_DIRNAME}/{platform}/{cat}/{date_name}/{fname}",
                                "force_download": ext.endswith((".pdf", ".txt", ".xlsx", ".xls", ".csv")),
                            })
                        days.append({"date": date_name, "files": files})

                categories.append({"category": cat, "days": days})

            groups.append({"group": platform, "categories": categories})
            continue

        # —— 其他平台（旧三层）：平台 / 日期 / 项目 / 文件 ——
        days = []
        for date_name in sorted(os.listdir(p_path), reverse=True):
            d_path = os.path.join(p_path, date_name)
            if not (DATE_RE.match(date_name) and os.path.isdir(d_path)):
                continue

            projects = []
            for proj in sorted(os.listdir(d_path)):
                proj_path = os.path.join(d_path, proj)
                if not os.path.isdir(proj_path):
                    continue

                files = []
                for fname in sorted(os.listdir(proj_path), key=file_sort_key):
                    fpath = os.path.join(proj_path, fname)
                    ext = fname.lower()
                    if not os.path.isfile(fpath):
                        continue
                    files.append({
                        "name": fname,
                        "url": f"{settings.DOWNLOAD_URL}{HISTORY_DIRNAME}/{platform}/{date_name}/{proj}/{fname}",
                        "force_download": ext.endswith((".pdf", ".txt", ".xlsx", ".xls", ".csv")),
                    })

                projects.append({"name": proj, "files": files})

            days.append({"date": date_name, "projects": projects})

        groups.append({"group": platform, "days": days})

    return render(request, "dashboard/file_download_history.html", {"groups": groups})


# === 新增：收集当日已取样过的实验号(match_sample)与条码(origin_barcode) ===
def _collect_today_sampled_codes() -> set[str]:
    """
    扫描 DOWNLOAD_ROOT 下“当日日期目录”内的所有 *.payload.json，
    提取 worksheet_table 里的 match_sample / origin_barcode，合并成一个大集合返回（统一大写、去空）。
    """
    root = Path(settings.DOWNLOAD_ROOT)
    if not root.exists():
        return set()

    today_str = timezone.localdate().strftime("%Y-%m-%d")
    sampled: set[str] = set()

    # 排除：历史文件、岗位清单
    excluded_dirs = {HISTORY_DIRNAME, STATION_DIRNAME}

    # 走全量递归：只要路径层级中包含 today_str，就认为是“当日日期子路径”
    for dirpath, dirnames, filenames in os.walk(root):
        p = Path(dirpath)

        # 快速跳过排除目录
        if any(part in excluded_dirs for part in p.parts):
            continue

        # 必须命中当日日期目录
        if today_str not in p.parts:
            continue

        for fn in filenames:
            lower = (fn or "").lower()
            if not lower.endswith(".payload.json"):
                continue

            fpath = p / fn
            try:
                payload = json.loads(fpath.read_text(encoding="utf-8"))
            except Exception:
                # 单个文件坏了不影响整体
                continue

            table = payload.get("worksheet_table") or []
            # worksheet_table: list[list[cell]]
            for row in table:
                if not isinstance(row, list):
                    continue
                for cell in row:
                    if not isinstance(cell, dict):
                        continue

                    ms = (cell.get("match_sample") or "").strip()
                    ob = (cell.get("origin_barcode") or "").strip()

                    if ms:
                        sampled.add(ms.upper())
                    if ob:
                        sampled.add(ob.upper())

    return sampled


# 给前端拉取当日集合（用于即时提示）
@require_GET
def file_replace_sampled_codes(request):
    """
    返回当日所有已取样过的条码/实验号集合（统一大写）。
    前端用于“新条码或实验号”的重复校验提示。
    """
    codes = sorted(_collect_today_sampled_codes())
    return JsonResponse({
        "date": timezone.localdate().strftime("%Y-%m-%d"),
        "codes": codes,
    })


@require_GET
def file_replace_get_payload(request):
    """
    根据上机列表文件名，返回对应的 payload.json 内容
    用于前端构建 match_sample -> origin_barcode 映射
    """
    filename = request.GET.get("filename", "").strip()
    if not filename:
        return JsonResponse({"ok": False, "message": "missing filename"}, status=400)
    
    # 推导 payload 文件名
    # X1_OnboardingList_...txt -> X1_WorkSheet_...payload.json
    if "OnboardingList" not in filename:
        return JsonResponse({"ok": False, "message": "not an onboarding file"}, status=400)
    
    # 替换文件名中的关键字段
    payload_filename = filename.replace("OnboardingList", "WorkSheet")
    
    # 去除原扩展名，添加 .payload.json
    base_name = os.path.splitext(payload_filename)[0]
    payload_filename = f"{base_name}.payload.json"
    
    # 在 DOWNLOAD_ROOT 下递归查找该文件（排除历史文件目录）
    root = settings.DOWNLOAD_ROOT
    payload_path = None
    
    for dirpath, dirnames, filenames in os.walk(root):
        # 跳过历史目录
        if HISTORY_DIRNAME in Path(dirpath).parts:
            continue
        
        if payload_filename in filenames:
            payload_path = os.path.join(dirpath, payload_filename)
            break
    
    if not payload_path or not os.path.exists(payload_path):
        return JsonResponse({
            "ok": False, 
            "message": f"payload file not found: {payload_filename}"
        }, status=404)
    
    # 读取并返回 payload 内容
    try:
        with open(payload_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        
        return JsonResponse({
            "ok": True,
            "payload": payload,
            "filename": payload_filename
        })
    except Exception as e:
        return JsonResponse({
            "ok": False,
            "message": f"failed to read payload: {str(e)}"
        }, status=500)



# 收集当前‘文件下载’页面中已有的上机列表文件名，用于后续匹配和替换
def _index_onboarding_files(root: str) -> dict:
    idx = {}
    hist_root = os.path.join(root, HISTORY_DIRNAME)

    for dirpath, _, filenames in os.walk(root):
        # 跳过历史目录
        if os.path.abspath(dirpath).startswith(os.path.abspath(hist_root) + os.sep):
            continue

        for fn in filenames:
            if "OnboardingList" in fn:
                idx[fn] = os.path.join(dirpath, fn)
    return idx

# 历史目录路径：把旧文件移动进去
def _history_path_for(abs_path: str, download_root: str) -> str:
    """
    将 DOWNLOAD_ROOT 下的文件 abs_path，映射到 DOWNLOAD_ROOT/历史文件/ 下的同相对路径。
    例如：
      DOWNLOAD_ROOT/NIMBUS/2025-12-29/25OHD/OnboardingList_xxx.txt
    -> DOWNLOAD_ROOT/历史文件/NIMBUS/2025-12-29/25OHD/OnboardingList_xxx.txt
    """
    rel = os.path.relpath(abs_path, download_root)
    return os.path.join(download_root, HISTORY_DIRNAME, rel)


# 从 OnboardingList 文件名推导 WorkSheet / payload 路径
def _derive_worksheet_names_from_onboarding(onboarding_name: str) -> tuple[str, str]:
    """
    输入：X1_OnboardingList_... .txt/.csv
    输出：
      worksheet_pdf_name: X1_WorkSheet_... .pdf
      worksheet_payload_name: X1_WorkSheet_... .payload.json
    """
    p = Path(onboarding_name)
    stem = p.stem
    # 统一按字段替换：OnboardingList -> WorkSheet
    ws_stem = stem.replace("OnboardingList", "WorkSheet")
    worksheet_pdf_name = f"{ws_stem}.pdf"
    worksheet_payload_name = f"{ws_stem}.payload.json"
    return worksheet_pdf_name, worksheet_payload_name


# 生成 Replace_ 文件名（与上机列表替换一致，防重名）
def _make_replace_name(dir_path: str, base_name: str) -> str:
    new_name = f"Replace_{base_name}"
    new_path = os.path.join(dir_path, new_name)
    if os.path.exists(new_path):
        ts = timezone.now().strftime("%Y%m%d_%H%M%S")
        stem, ext = os.path.splitext(new_name)
        new_name = f"{stem}_{ts}{ext}"
    return new_name

# 新增工具函数：条码解析（更新 cut/sub/origin）
_BARCODE_SPLIT_RE = re.compile(r"^(.+?)(-\w+)?$")

def _parse_barcode(new_code: str) -> tuple[str, str, str]:
    """
    输入: 2437871821-01 or 6810526380 or NOTUBE
    输出: (cut_barcode, sub_barcode, origin_barcode)
    """
    s = (new_code or "").strip()
    if not s:
        return "", "", ""
    if s.upper() == "NOTUBE":
        return "NOTUBE", "", "NOTUBE"

    m = _BARCODE_SPLIT_RE.match(s)
    if not m:
        return s, "", s
    cut = m.group(1) or ""
    sub = m.group(2) or ""
    origin = f"{cut}{sub}" if sub else cut
    return cut, sub, origin

# 新增工具函数：在 payload 里定位 cell，并写入 highlight
def _build_cell_index(payload: dict) -> dict:
    """
    返回 well_str -> cell dict 的索引
    """
    idx = {}
    for row in payload.get("worksheet_table", []):
        for cell in row:
            ws = (cell.get("well_str") or "").strip().upper()
            if ws:
                idx[ws] = cell
    return idx

def _mark_cell_highlight(cell: dict):
    cell["highlight"] = True


# 新增核心函数：对 payload 应用 used / nouse / delete，并保持 error_rows 不动
_SAMPLE_RE = re.compile(r"[A-Za-z]")  # 含字母则更像“实验号/样本名”

def _apply_replace_to_payload(payload: dict, replace_reason: str, entries: list[dict]) -> tuple[dict, list[str]]:
    """
    返回：修改后的 payload + 高亮孔位列表
    注意：不改 error_rows（按你的需求“报错信息表不变”）
    """
    cell_index = _build_cell_index(payload)
    highlighted = []

    def get_cell_by_vialpos(vp: str):
        norm = _normalize_user_vialpos(vp)
        if norm.get("ok"):
            return cell_index.get(norm["well"])
        return None

    for e in entries:
        vp = e.get("vialpos", "")
        cell = get_cell_by_vialpos(vp)

        if not cell:
            # 兜底：允许通过 old/code 在 payload 里匹配（可选）
            key = (e.get("old") or e.get("code") or "").strip()
            if key:
                for c in cell_index.values():
                    if key == (c.get("match_sample") or "").strip():
                        cell = c; break
                    if key == (c.get("origin_barcode") or "").strip():
                        cell = c; break
                    if key == (c.get("cut_barcode") or "").strip():
                        cell = c; break

        if not cell:
            raise ValueError(f"未在工作清单payload中定位到孔位：{vp}")

        # ===== 应用规则 =====
        if replace_reason == "used":
            new_val = (e.get("new") or "").strip()
            old_val = (e.get("old") or "").strip()

            # old 为空时，根据 new 判断更新实验号还是条码
            if _SAMPLE_RE.search(new_val):  # 含字母：当作实验号
                cell["match_sample"] = new_val
            else:  # 当作条码
                cut, sub, origin = _parse_barcode(new_val)
                cell["cut_barcode"] = cut
                cell["sub_barcode"] = sub
                cell["origin_barcode"] = origin

            # 你需求“孔位/条码/实验号替换”，因此即使 old 填的是实验号/条码，都以 new 覆盖显示
            # （这里不强制校验 old 是否与 cell 现值一致，避免用户只按孔位替换时失败）
            _mark_cell_highlight(cell)

        elif replace_reason == "nouse":
            new_val = (e.get("new") or "").strip()
            if _SAMPLE_RE.search(new_val):
                cell["match_sample"] = new_val
            else:
                cut, sub, origin = _parse_barcode(new_val)
                cell["cut_barcode"] = cut
                cell["sub_barcode"] = sub
                cell["origin_barcode"] = origin
            _mark_cell_highlight(cell)

        elif replace_reason == "delete":
            # 删除：恢复 NOTUBE / No match 显示
            cell["match_sample"] = "No match"
            cell["cut_barcode"] = "NOTUBE"
            cell["sub_barcode"] = ""
            cell["origin_barcode"] = "NOTUBE"
            _mark_cell_highlight(cell)

        else:
            raise ValueError(f"未知 replace_reason: {replace_reason}")

        highlighted.append((cell.get("well_str") or "").strip())

    return payload, highlighted



# 新增核心函数：渲染 export_pdf.html 并写 PDF（复用你现有 export_files 的 CSS/Font 配置）
def _render_payload_to_pdf(payload: dict, out_pdf_path: str):
    pdf_html = render_to_string("dashboard/export_pdf.html", payload)

    font_config = FontConfiguration()
    pdf_css = CSS(string="""
        @page { size: A4; margin: 10mm; }
        body { font-family: "Noto Sans CJK SC", "SimSun", sans-serif; font-size: 10px; }
    """, font_config=font_config)

    HTML(string=pdf_html).write_pdf(
        out_pdf_path,
        stylesheets=[pdf_css],
        font_config=font_config
    )

# 新增：WorkSheet 替换总函数（找旧文件 → 生成 Replace_ 新文件 → 旧文件进历史）
def _replace_worksheet_after_onboarding_replace(root: str, onboarding_uploaded_name: str,
                                              replace_reason: str, entries: list[dict], target_dir: str):
    """
    root: DOWNLOAD_ROOT
    target_dir: 上机列表所在目录（也就是 worksheet/pdf/payload 所在目录）
    """
    ws_pdf_name, ws_payload_name = _derive_worksheet_names_from_onboarding(onboarding_uploaded_name)

    ws_pdf_path = os.path.join(target_dir, ws_pdf_name)
    ws_payload_path = os.path.join(target_dir, ws_payload_name)

    if not os.path.exists(ws_pdf_path):
        raise FileNotFoundError(f"未找到对应工作清单PDF：{ws_pdf_name}")
    if not os.path.exists(ws_payload_path):
        raise FileNotFoundError(f"未找到对应payload.json：{ws_payload_name}")

    # 1) 读取 payload
    with open(ws_payload_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    # 2) 修改 payload + 标红
    payload, _ = _apply_replace_to_payload(payload, replace_reason, entries)

    # 3) 生成 Replace_ WorkSheet PDF
    new_ws_pdf_name = _make_replace_name(target_dir, ws_pdf_name)
    new_ws_pdf_path = os.path.join(target_dir, new_ws_pdf_name)

    _render_payload_to_pdf(payload, new_ws_pdf_path)

    # 4) 同步生成 Replace_ payload.json（与PDF同 stem）
    new_payload_name = f"{Path(new_ws_pdf_name).stem}.payload.json"
    _dump_payload_json(target_dir, payload, filename=new_payload_name)  # 你 views 里已有此函数

    # 5) 旧 PDF + 旧 payload 进历史目录
    hist_pdf_path = _history_path_for(ws_pdf_path, root)
    os.makedirs(os.path.dirname(hist_pdf_path), exist_ok=True)
    os.replace(ws_pdf_path, hist_pdf_path)

    hist_payload_path = _history_path_for(ws_payload_path, root)
    os.makedirs(os.path.dirname(hist_payload_path), exist_ok=True)
    os.replace(ws_payload_path, hist_payload_path)


# 解析分隔符 + 读取表格（只要能稳定分列即可）
def _guess_delimiter(line: str) -> str:
    tab = line.count("\t")
    comma = line.count(",")
    semi = line.count(";")
    if tab >= comma and tab >= semi:
        return "\t"
    if comma >= semi:
        return ","
    return ";"

def _read_text_with_fallback(path: str) -> tuple[str, str]:
    """
    以二进制读取后按多编码尝试解码，返回 (text, encoding_used)
    兼容：utf-8 / utf-8-sig / gb18030 / gbk
    """
    with open(path, "rb") as f:
        raw = f.read()

    last_err = None
    for enc in ("gbk","utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError as e:
            last_err = e

    # 最后兜底：不阻断，但会出现 replacement char
    return raw.decode("utf-8", errors="replace"), "utf-8(replace)"


# 识别：哪些列在“整张表”范围内与第 1 列（index 0）逐行完全相同
def _detect_columns_equal_to_first(rows: list[list[str]]) -> list[int]:
    """
    rows: 解析后的二维表（含表头/数据行都可以；建议你传“数据行”）
    返回：所有满足“每一行 col[i] == col[0]”的列索引 i（i>=1）
    
    判定规则（尽量稳妥）：
    - 若某行缺少该列，视为“不相同”（直接淘汰该列）
    - 比较时做 strip()，并将 None 视为空串
    """
    if not rows:
        return []

    # 计算最大列数
    max_cols = max(len(r) for r in rows if r)
    if max_cols <= 1:
        return []

    candidates = list(range(1, max_cols))  # 只看第2列起
    keep = []

    for ci in candidates:
        ok = True
        for r in rows:
            if len(r) <= ci:
                ok = False
                break
            v0 = (r[0] or "").strip()
            vi = (r[ci] or "").strip()
            if vi != v0:
                ok = False
                break
        if ok:
            keep.append(ci)

    return keep


# 孔位清洗逻辑（与前端一致：只取最后一段）
_SPLIT_VIAL_RE = re.compile(r"[\s:\-_]+")
def _clean_vialpos(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    parts = [p for p in _SPLIT_VIAL_RE.split(s) if p]
    return parts[-1] if parts else s

# 把用户输入的孔位规范化（支持 A1(1) / A1 / a1 / 1）
_WELL_RE = re.compile(r"^([A-H])\s*([1-9]|1[0-2])$", re.I)
def _num_to_well(n: int) -> str | None:
    if not (1 <= n <= 96):
        return None
    row = chr(ord("A") + (n - 1) // 12)
    col = (n - 1) % 12 + 1
    return f"{row}{col}"

def _well_to_num(well: str) -> int | None:
    m = _WELL_RE.match((well or "").strip())
    if not m:
        return None
    row = m.group(1).upper()
    col = int(m.group(2))
    row_idx = ord(row) - ord("A")
    return row_idx * 12 + col

def _normalize_user_vialpos(s: str) -> dict:
    """
    返回：
      {ok: True, well: "A1", num: 1}
    支持输入：A1(1)、A1、a1、1、01
    """
    raw = (s or "").strip()
    if not raw:
        return {"ok": False}

    raw = raw.replace("（", "(").replace("）", ")").strip()

    paren_num = None
    pm = re.match(r"^(.*)\(\s*(\d+)\s*\)\s*$", raw)
    if pm:
        raw = (pm.group(1) or "").strip()
        paren_num = int(pm.group(2))

    # 纯数字
    if re.fullmatch(r"\d+", raw):
        n = int(raw)
        well = _num_to_well(n)
        if not well:
            return {"ok": False}
        if paren_num is not None and paren_num != n:
            return {"ok": False}
        return {"ok": True, "well": well, "num": n}

    # A1-H12
    n = _well_to_num(raw.upper())
    if not n:
        return {"ok": False}
    if paren_num is not None and paren_num != n:
        return {"ok": False}
    return {"ok": True, "well": _num_to_well(n), "num": n}


def file_replace(request):
    if request.method == "POST":
        upload = request.FILES.get("replace_file")
        if not upload:
            return render(request, "dashboard/error.html", {
                "message": "未检测到上传文件，请先选择需要替换的上机列表文件。"
            })

        uploaded_name = Path(upload.name).name
        project_name = _extract_project_from_onboarding_filename(uploaded_name)

        root = settings.DOWNLOAD_ROOT
        os.makedirs(root, exist_ok=True)

        # 建立“已有上机列表文件名索引”
        exist_map = _index_onboarding_files(root)  # 文件名包含 OnboardingList :contentReference[oaicite:6]{index=6}
        target_path = exist_map.get(uploaded_name)
        if not target_path:
            return render(request, "dashboard/error.html", {
                "message": f"上传文件名【{uploaded_name}】在“文件下载”中未找到同名上机列表文件，请确认文件名必须完全一致。"
            })

        replace_reason = (request.POST.get("replace_reason") or "").strip()

        # ===== 只先实现：已用孔位替换 used =====
        if replace_reason == "used":
            used_vialpos = request.POST.getlist("used_vialpos[]")
            used_old     = request.POST.getlist("used_old_barcode[]")
            used_new     = request.POST.getlist("used_new_barcode[]")

            # 基本校验：三列长度一致
            if not (len(used_vialpos) == len(used_old) == len(used_new)):
                return render(request, "dashboard/error.html", {
                    "message": "提交数据不完整：孔位/旧条码/新条码行数不一致。"
                })

            # 过滤空行（允许用户多加空行）
            entries = []
            for vp, old, new in zip(used_vialpos, used_old, used_new):
                vp = (vp or "").strip()
                old = (old or "").strip()
                new = (new or "").strip()
                if not (vp or old or new):
                    continue
                if not new:
                    return render(request, "dashboard/error.html", {
                        "message": "存在未填写的新条码/实验号，请补全后再提交。"
                    })
                entries.append({"vialpos": vp, "old": old, "new": new})

            if not entries:
                return render(request, "dashboard/error.html", {
                    "message": "未检测到任何有效替换行，请填写后再提交。"
                })

            # 读取目标文件
            with open(target_path, "r", encoding="utf-8", errors="replace") as f:
                text, file_enc = _read_text_with_fallback(target_path)

            lines = [l for l in text.splitlines() if l.strip() != ""]
            if len(lines) < 2:
                return render(request, "dashboard/error.html", {
                    "message": f"上机列表文件内容不足，无法替换：{uploaded_name}"
                })

            delimiter = _guess_delimiter(lines[0])
            header = lines[0].split(delimiter)

            # 找 VialPos 列（匹配 'vialpos' 或 'vial position'，大小写不敏感）
            def _norm_h(h: str) -> str:
                return (h or "").strip().lower().replace("% header=", "").strip()

            headers_norm = [_norm_h(h) for h in header]
            vial_idx = -1
            for i, h in enumerate(headers_norm):
                if h == "vialpos" or h == "vial position":
                    vial_idx = i
                    break
                # ✅ 新增：中文列名（需要原始表头匹配，不能用 lower()）
                if header[i].strip() == "样品瓶":
                    vial_idx = i
                    break

            if vial_idx == -1:
                return render(request, "dashboard/error.html", {
                    "message": f"未找到孔位列（VialPos / Vial position），无法替换：{uploaded_name}"
                })

            rows = []
            for i in range(1, len(lines)):
                cols = lines[i].split(delimiter)
                # 补齐列数，避免越界
                if len(cols) < len(header):
                    cols = cols + [""] * (len(header) - len(cols))
                rows.append(cols)

            # ✅ 新增:提取进样盘号前缀(从第一个有效孔位中提取)
            drawer_prefix = ""  # 进样盘号前缀,如 "Drawer 2:Slot1:"
            for cols in rows:
                vp_raw = (cols[vial_idx] or "").strip()
                if not vp_raw or vp_raw.upper() == "NOTUBE":
                    continue

                # 使用与 _clean_vialpos 相同的分隔符规则 [\s:\-_]+
                parts = [p for p in _SPLIT_VIAL_RE.split(vp_raw) if p]
                if len(parts) >= 2:
                    # 提取除最后一段外的所有部分作为前缀
                    # 例如: "Drawer 2:Slot1:86" -> parts=["Drawer", "2", "Slot1", "86"]
                    # 前缀部分: "Drawer 2:Slot1:"
                    prefix_parts = parts[:-1]
                    # 重建前缀(使用原始字符串中的分隔符)
                    last_part = parts[-1]
                    idx = vp_raw.rfind(last_part)
                    if idx > 0:
                        drawer_prefix = vp_raw[:idx]  # 包含尾部分隔符
                        break

            # ===== ✅ 新增：检测“与第一列逐行完全相同”的列，替换时需要联动更新 =====
            same_as_first_cols = _detect_columns_equal_to_first(rows)

            # 执行替换：仅替换第一列（index=0）
            def _row_match(cols: list[str], entry: dict) -> bool:
                # 1) 优先用 old 匹配第一列（如果 old 有填）
                old = entry["old"]
                if old and (cols[0] or "").strip() == old:
                    return True

                # 2) 用 vialpos 匹配 VialPos 列（支持 A1(1)/A1/a1/1）
                vp_in = entry["vialpos"]
                if not vp_in:
                    return False

                norm = _normalize_user_vialpos(vp_in)
                vp_cell_raw = (cols[vial_idx] or "").strip()
                vp_clean = _clean_vialpos(vp_cell_raw)

                # 兼容：文件里可能是 "A1" / "1" / "Rack-...-A1"
                vp_clean_up = vp_clean.upper()

                if norm.get("ok"):
                    if vp_clean_up == norm["well"]:
                        return True
                    if vp_clean == str(norm["num"]):
                        return True

                # 兜底：如果用户输入本身就是 "A1" 或 "1"，也直接比一次
                if vp_clean_up == vp_in.strip().upper():
                    return True
                if vp_clean == vp_in.strip():
                    return True

                return False

            replaced = 0
            for entry in entries:
                hit = False
                for cols in rows:
                    if _row_match(cols, entry):

                        # ✅ 统一处理：将新值转为文本格式（防止Excel识别为数值）
                        new_code_text = str(entry["new"]).strip()
                        if new_code_text.isdigit():
                            new_code_formatted = f"\t{new_code_text}"  # 制表符前缀强制文本格式
                        else:
                            new_code_formatted = new_code_text

                        # 1) 永远替换第一列
                        cols[0] = new_code_formatted

                        # 2) ✅ 联动替换"与第一列逐行完全相同"的列（统一使用格式化后的值）
                        for j in same_as_first_cols:
                            if 0 <= j < len(cols):
                                cols[j] = new_code_formatted

                        # 3) ✅ 新增：如果提取到进样盘号前缀,且当前孔位列没有进样盘号,则拼接
                        if drawer_prefix:
                            vp_current = (cols[vial_idx] or "").strip()
                            if vp_current and drawer_prefix:
                                # 判断当前孔位是否已包含分隔符(任意一种: 空格/冒号/短横线/下划线)
                                has_prefix = bool(_SPLIT_VIAL_RE.search(vp_current))
                                
                                if not has_prefix:
                                    # 当前孔位无前缀,需要拼接
                                    norm = _normalize_user_vialpos(vp_current)
                                    if norm.get("ok"):
                                        # 优先写成 1..96 数字形式,并拼接进样盘号前缀
                                        cols[vial_idx] = f"{drawer_prefix}{norm['num']}"
                                    else:
                                        # 兜底:直接拼接用户输入
                                        cols[vial_idx] = f"{drawer_prefix}{vp_current}"

                        replaced += 1
                        hit = True
                        break
                if not hit:
                    return render(request, "dashboard/error.html", {
                        "message": f"未找到可替换行：孔位={entry['vialpos']} / 当前条码={entry['old']}"
                    })


        # ===== 未用孔位替换 nouse =====
        elif replace_reason == "nouse":
            # ===== ★ 新增：定位孔识别函数 =====
            def _is_locator_row(cols: list[str], first_col_idx: int = 0) -> bool:
                """
                判断某一行是否为定位孔行
                规则：第一列内容包含'X'关键词，且字符串总长度 <= 3
                """
                first_val = (cols[first_col_idx] or "").strip().upper()
                if 'X' in first_val and len(first_val) <= 3:
                    return True
                return False
            
            def _find_locator_row_index(rows: list[list[str]]) -> int:
                """
                在 rows 列表中查找定位孔行的索引
                返回：定位孔行的索引（从0开始），若未找到返回 -1
                """
                for idx, cols in enumerate(rows):
                    if _is_locator_row(cols):
                        return idx
                return -1
    
            nouse_new = request.POST.getlist("nouse_new_barcode[]")
            nouse_vp  = request.POST.getlist("nouse_vialpos[]")

            if len(nouse_new) != len(nouse_vp):
                return render(request, "dashboard/error.html", {
                    "message": "提交数据不完整：新条码/孔位行数不一致。"
                })

            entries = []
            for new_code, vp in zip(nouse_new, nouse_vp):
                new_code = (new_code or "").strip()
                vp = (vp or "").strip()
                if not (new_code or vp):
                    continue
                if not new_code:
                    return render(request, "dashboard/error.html", {
                        "message": "存在未填写的新条码/实验号，请补全后再提交。"
                    })
                if not vp:
                    return render(request, "dashboard/error.html", {
                        "message": "存在未填写的孔位（VialPos），请补全后再提交。"
                    })
                entries.append({"new": new_code, "vialpos": vp})

            if not entries:
                return render(request, "dashboard/error.html", {
                    "message": "未检测到任何有效替换行，请填写后再提交。"
                })


            # 读取目标文件
            with open(target_path, "r", encoding="utf-8", errors="replace") as f:
                text, file_enc = _read_text_with_fallback(target_path)

            lines = [l for l in text.splitlines() if l.strip() != ""]
            if len(lines) < 2:
                return render(request, "dashboard/error.html", {
                    "message": f"上机列表文件内容不足，无法替换：{uploaded_name}"
                })

            delimiter = _guess_delimiter(lines[0])
            header = lines[0].split(delimiter)

            def _norm_h(h: str) -> str:
                return (h or "").strip().lower().replace("% header=", "").strip()

            headers_norm = [_norm_h(h) for h in header]

            vial_idx = -1
            for i, h in enumerate(headers_norm):
                if h == "vialpos" or h == "vial position":
                    vial_idx = i
                    break
                # ✅ 新增：中文列名（需要原始表头匹配，不能用 lower()）
                if header[i].strip() == "样品瓶":
                    vial_idx = i
                    break

            if vial_idx == -1:
                return render(request, "dashboard/error.html", {
                    "message": f"未找到孔位列（VialPos / Vial position / 样品瓶），无法替换：{uploaded_name}。当前表头：{' | '.join(header)}"
                })

            rows = []
            for i in range(1, len(lines)):
                cols = lines[i].split(delimiter)
                if len(cols) < len(header):
                    cols = cols + [""] * (len(header) - len(cols))
                rows.append(cols)

            # ✅ 新增:提取进样盘号前缀(从第一个有效孔位中提取)
            drawer_prefix = ""  # 进样盘号前缀,如 "Drawer 2:Slot1:"
            for cols in rows:
                vp_raw = (cols[vial_idx] or "").strip()
                if not vp_raw or vp_raw.upper() == "NOTUBE":
                    continue
                
                # 使用与 _clean_vialpos 相同的分隔符规则 [\s:\-_]+
                parts = [p for p in _SPLIT_VIAL_RE.split(vp_raw) if p]
                if len(parts) >= 2:
                    # 提取除最后一段外的所有部分作为前缀
                    # 例如: "Drawer 2:Slot1:86" -> parts=["Drawer", "2", "Slot1", "86"]
                    # 前缀部分: "Drawer 2:Slot1:"
                    prefix_parts = parts[:-1]
                    # 重建前缀(使用原始字符串中的分隔符)
                    last_part = parts[-1]
                    idx = vp_raw.rfind(last_part)
                    if idx > 0:
                        drawer_prefix = vp_raw[:idx]  # 包含尾部分隔符
                        break

            # ===== ✅ 新增：检测“与第一列逐行完全相同”的列，替换时需要联动更新 =====
            same_as_first_cols = _detect_columns_equal_to_first(rows)

            # 找一条“临床样品行”作为模板：第一列不是 NOTUBE 即可
            template_row = None
            for cols in rows:
                first = (cols[0] or "").strip().upper()
                if first and first != "NOTUBE":
                    template_row = cols
                    break

            if template_row is None:
                return render(request, "dashboard/error.html", {
                    "message": "未找到可用于复制的模板行（第一列非 NOTUBE），无法执行未用孔位替换。"
                })


            # 根据用户输入孔位定位行
            def _match_row_by_vialpos(cols: list[str], vp_in: str) -> bool:
                vp_in = (vp_in or "").strip()
                if not vp_in:
                    return False

                norm = _normalize_user_vialpos(vp_in)
                vp_cell_raw = (cols[vial_idx] or "").strip()
                vp_clean = _clean_vialpos(vp_cell_raw)
                vp_clean_up = vp_clean.upper()

                if norm.get("ok"):
                    if vp_clean_up == norm["well"]:
                        return True
                    if vp_clean == str(norm["num"]):
                        return True

                # 兜底直接比
                if vp_clean_up == vp_in.upper():
                    return True
                if vp_clean == vp_in:
                    return True
                return False

            replaced = 0
            used_vps_seen = set()

            for entry in entries:
                vp_in = entry["vialpos"]
                # 防止重复孔位（同一次提交）
                vp_key = vp_in.strip().upper()
                if vp_key in used_vps_seen:
                    return render(request, "dashboard/error.html", {
                        "message": f"提交中存在重复孔位：{vp_in}"
                    })
                used_vps_seen.add(vp_key)

                hit_cols = None
                for cols in rows:
                    if _match_row_by_vialpos(cols, vp_in):
                        hit_cols = cols
                        break

                # ★ 修改：若文件中缺失该孔位行，则创建新行并插入到定位孔下一行
                if hit_cols is None:
                    # 复制一行模板，保证列数一致
                    hit_cols = list(template_row)

                    # 规范化孔位写入：优先写成 1..96 数字
                    norm = _normalize_user_vialpos(vp_in)  
                    if norm.get("ok"):
                        # 优先写成 1..96 数字形式,并拼接进样盘号前缀
                        vialpos_value = f"{drawer_prefix}{norm['num']}" if drawer_prefix else str(norm["num"])
                    else:
                        # 兜底:直接写用户输入,拼接进样盘号前缀
                        vialpos_value = f"{drawer_prefix}{vp_in.strip()}" if drawer_prefix else vp_in.strip()

                    hit_cols[vial_idx] = vialpos_value

                    # 第一列先填 NOTUBE 占位（可选），后面会被 hit_cols[0]=entry["new"] 覆盖
                    if len(hit_cols) > 0 and not (hit_cols[0] or "").strip():
                        hit_cols[0] = "NOTUBE"

                    # ===== ★ 新增：按定位孔位置插入 =====
                    locator_idx = _find_locator_row_index(rows)
                    if locator_idx != -1:
                        # 找到定位孔：插入到定位孔的下一行（索引 +1）
                        rows.insert(locator_idx + 1, hit_cols)
                    else:
                        # 未找到定位孔：兜底追加到末尾（保持原有行为）
                        rows.append(hit_cols)

                # 执行替换：除第一列和孔位列之外，其他列复制模板行；再覆盖第一列为新条码；孔位列保持原值
                old_vp_val = hit_cols[vial_idx]
                for j in range(len(hit_cols)):
                    if j == 0 or j == vial_idx:
                        continue
                    hit_cols[j] = template_row[j]

                # ✅ 统一处理：将新值转为文本格式（防止Excel识别为数值）
                new_code_text = str(entry["new"]).strip()
                if new_code_text.isdigit():
                    new_code_formatted = f"\t{new_code_text}"  # 制表符前缀强制文本格式
                else:
                    new_code_formatted = new_code_text

                # 替换第一列
                hit_cols[0] = new_code_formatted

                # ✅ 联动替换"与第一列逐行完全相同"的列（统一使用格式化后的值）
                for j in same_as_first_cols:
                    if 0 <= j < len(hit_cols):
                        hit_cols[j] = new_code_formatted

                hit_cols[vial_idx] = old_vp_val

                replaced += 1

        
        # ===== 孔位删除 delete =====
        elif replace_reason == "delete":
            del_vialpos = request.POST.getlist("delete_vialpos[]")
            del_code    = request.POST.getlist("delete_barcode[]")

            if len(del_vialpos) != len(del_code):
                return render(request, "dashboard/error.html", {
                    "message": "提交数据不完整：孔位/条码行数不一致。"
                })

            entries = []
            for vp, code in zip(del_vialpos, del_code):
                vp = (vp or "").strip()
                code = (code or "").strip()

                # 允许其中一列为空（前端会自动回填），但至少要有一个
                if not (vp or code):
                    continue

                entries.append({"vialpos": vp, "code": code})

            if not entries:
                return render(request, "dashboard/error.html", {
                    "message": "未检测到任何有效删除行，请填写后再提交。"
                })

            # 读取目标文件
            with open(target_path, "r", encoding="utf-8", errors="replace") as f:
                text, file_enc = _read_text_with_fallback(target_path)

            lines = [l for l in text.splitlines() if l.strip() != ""]
            if len(lines) < 2:
                return render(request, "dashboard/error.html", {
                    "message": f"上机列表文件内容不足，无法删除：{uploaded_name}"
                })

            delimiter = _guess_delimiter(lines[0])
            header = lines[0].split(delimiter)

            def _norm_h(h: str) -> str:
                return (h or "").strip().lower().replace("% header=", "").strip()

            headers_norm = [_norm_h(h) for h in header]

            vial_idx = -1
            for i, h in enumerate(headers_norm):
                if h == "vialpos" or h == "vial position":
                    vial_idx = i
                    break
                # ✅ 新增：中文列名（需要原始表头匹配，不能用 lower()）
                if header[i].strip() == "样品瓶":
                    vial_idx = i
                    break

            if vial_idx == -1:
                return render(request, "dashboard/error.html", {
                    "message": f"未找到孔位列（VialPos / Vial position / 样品瓶），无法删除：{uploaded_name}。当前表头：{' | '.join(header)}"
                })

            # rows：保持列数与 header 一致
            rows = []
            for i in range(1, len(lines)):
                cols = lines[i].split(delimiter)
                if len(cols) < len(header):
                    cols = cols + [""] * (len(header) - len(cols))
                rows.append(cols)

            # 建索引：barcode -> row
            # 第一列为 barcode/实验号列（你们系统约定）
            barcode_to_row = {}
            for cols in rows:
                code0 = (cols[0] or "").strip()
                if code0:
                    barcode_to_row[code0] = cols

            def _match_row_by_vialpos(cols: list[str], vp_in: str) -> bool:
                vp_in = (vp_in or "").strip()
                if not vp_in:
                    return False

                norm = _normalize_user_vialpos(vp_in)
                vp_cell_raw = (cols[vial_idx] or "").strip()
                vp_clean = _clean_vialpos(vp_cell_raw)
                vp_clean_up = vp_clean.upper()

                if norm.get("ok"):
                    if vp_clean_up == norm["well"]:
                        return True
                    if vp_clean == str(norm["num"]):
                        return True

                # 兜底直接比
                if vp_clean_up == vp_in.upper():
                    return True
                if vp_clean == vp_in:
                    return True
                return False

            # 防重复（同一提交）
            seen_keys = set()
            deleted = 0

            # ✅ 新增：收集要删除的行对象（而不是修改内容）
            rows_to_delete = []

            for entry in entries:
                vp_in = entry["vialpos"]
                code_in = entry["code"]

                key = (vp_in.strip().upper() if vp_in else "") + "|" + (code_in.strip())
                if key in seen_keys:
                    return render(request, "dashboard/error.html", {
                        "message": f"提交中存在重复删除行：孔位={vp_in}, 条码/实验号={code_in}"
                    })
                seen_keys.add(key)

                hit_cols = None

                # 优先用 code 命中（更快）
                if code_in:
                    hit_cols = barcode_to_row.get(code_in)

                # 若没命中，再用孔位命中
                if hit_cols is None and vp_in:
                    for cols in rows:
                        if _match_row_by_vialpos(cols, vp_in):
                            hit_cols = cols
                            break

                if hit_cols is None:
                    return render(request, "dashboard/error.html", {
                        "message": f"未在上机列表中找到要删除的记录：孔位={vp_in} 条码/实验号={code_in}"
                    })

                # 若两列都填了，校验一致性（防误删）
                if vp_in:
                    ok = _match_row_by_vialpos(hit_cols, vp_in)
                    if not ok:
                        return render(request, "dashboard/error.html", {
                            "message": f"孔位与条码/实验号不匹配：孔位={vp_in} 条码/实验号={code_in}"
                        })

                if code_in:
                    code0 = (hit_cols[0] or "").strip()
                    if code0 != code_in:
                        return render(request, "dashboard/error.html", {
                            "message": f"孔位与条码/实验号不匹配：孔位={vp_in} 条码/实验号={code_in}"
                        })

                # ✅ 修改：不再清空内容，而是标记该行对象待删除
                rows_to_delete.append(hit_cols)
                deleted += 1

            # ✅ 新增：从 rows 列表中移除标记的行
            for row in rows_to_delete:
                if row in rows:
                    rows.remove(row)

        else:
            pass
        
        # 生成新文件名：Replace_ + 原文件名（防重名）
        dir_path = os.path.dirname(target_path)
        base = os.path.basename(target_path)
        new_name = f"Replace_{base}"
        new_path = os.path.join(dir_path, new_name)

        if os.path.exists(new_path):
            ts = timezone.now().strftime("%Y%m%d_%H%M%S")
            stem, ext = os.path.splitext(new_name)
            new_name = f"{stem}_{ts}{ext}"
            new_path = os.path.join(dir_path, new_name)

        # 写入新文件（保持原分隔符）
        out_lines = []
        out_lines.append(delimiter.join(header))
        for cols in rows:
            out_lines.append(delimiter.join(cols))
        out_text = "\n".join(out_lines) + "\n"

        with open(new_path, "w", encoding="utf-8") as f:
            f.write(out_text)
            
        # 把旧文件移动到历史目录
        hist_path = _history_path_for(target_path, root)
        os.makedirs(os.path.dirname(hist_path), exist_ok=True)
        os.replace(target_path, hist_path)

        # ===== 新增：同步替换 WorkSheet PDF（工作清单）=====
        try:
            # 说明：
            # - uploaded_name：用户上传的上机列表文件名（OnboardingList...）
            # - replace_reason：used / nouse / delete
            # - entries：三种分支都构造了 entries（孔位/旧/新）
            # - dir_path：公共出口里已经有 dir_path = os.path.dirname(target_path)
            _replace_worksheet_after_onboarding_replace(
                root=root,
                onboarding_uploaded_name=uploaded_name,
                replace_reason=replace_reason,
                entries=entries,
                target_dir=dir_path,
            )
        except Exception as e:
            return render(request, "dashboard/error.html", {
                "message": f"上机列表替换成功，但工作清单替换失败：{str(e)}"
            })

        # ✅ 完成：下载目录里将只剩 Replace_ 文件（新文件名出现在文件下载页）
        # 建议直接回到文件下载页
        return redirect("file_download")


    # GET：打开页面时，把“文件下载页已有的上机列表文件名”传给前端
    root = settings.DOWNLOAD_ROOT
    os.makedirs(root, exist_ok=True)
    exist_map = _index_onboarding_files(root)
    onboarding_filenames = sorted(exist_map.keys())

    return render(request, "dashboard/file_replace.html", {
        "onboarding_filenames": onboarding_filenames
    })


# 3 后台参数配置
def project_config(request):
    project_configs = SamplingConfiguration.objects.all().order_by('project_name')
    return render(request, 'dashboard/config/project_config.html', {
        'project_configs': project_configs
    })

def project_config_create(request):
    if request.method == 'POST':
        # 1️⃣ 获取表单字段
        project_name = request.POST.get('project_name', '').strip()
        project_name_full = request.POST.get('project_name_full', '').strip()
        sampling_method = request.POST.get('sampling_method', '').strip()
        curve_points = request.POST.get('curve_points', '').strip()
        qc_groups = request.POST.get('qc_groups', '').strip()
        qc_levels = request.POST.get('qc_levels', '').strip()
        qc_insert = request.POST.get('qc_insert', '').strip()
        test_count = request.POST.get('test_count', '').strip()
        layout = request.POST.get('layout', '').strip()
        default_upload_instrument = request.POST.get('default_upload_instrument', '').strip()
        systerm_num = request.POST.get('systerm_num', '').strip()
        mapping_file = request.FILES.get('mapping_file')

        # 2️⃣ 校验是否已存在相同配置
        duplicate_qs = SamplingConfiguration.objects.filter(
            project_name=project_name,
            default_upload_instrument=default_upload_instrument,
            systerm_num=systerm_num
        )
        if duplicate_qs.exists():
            return render(request, 'dashboard/error.html', {
                "message": f"项目 [{project_name}] 已存在相同的仪器编号 [{default_upload_instrument}] 与系统号 [{systerm_num}] 配置，请勿重复创建。"
            })

        # 3️⃣ 创建新配置
        instance = SamplingConfiguration(
            project_name=project_name,
            project_name_full=project_name_full,
            sampling_method=sampling_method,
            curve_points=curve_points,
            qc_groups=qc_groups,
            qc_levels=qc_levels,
            qc_insert=qc_insert,
            test_count=test_count,
            layout=layout,
            default_upload_instrument=default_upload_instrument,
            systerm_num=systerm_num,
            mapping_file=mapping_file,
        )
        instance.save()
        return redirect('project_config')

    else:
        curve_range = list(range(6, 11))  # [6, 7, 8, 9, 10]
        return render(request, 'dashboard/config/project_config_create.html', {
            'curve_range': curve_range
        })

def project_config_view(request, pk):
    config = get_object_or_404(SamplingConfiguration, pk=pk)
    return render(request, 'dashboard/config/project_config_view.html', {
        'config': config
    })

def project_config_edit(request, pk):
    config = get_object_or_404(SamplingConfiguration, pk=pk)
    if request.method == 'POST':
        # 保存旧文件引用
        old_mapping_file = config.mapping_file.path if config.mapping_file else None

        form = SamplingConfigurationForm(request.POST, request.FILES, instance=config)
        if form.is_valid():
            # 检查是否上传新文件，如果上传了，删除原文件
            if 'mapping_file' in request.FILES and old_mapping_file and os.path.exists(old_mapping_file):
                os.remove(old_mapping_file)

            form.save()
            return redirect('project_config')
    else:
        curve_range = list(range(6, 11)) 
        form = SamplingConfigurationForm(instance=config)
    return render(request, 'dashboard/config/project_config_edit.html', {
        'form': form,
        'config': config,
    })

def project_config_delete(request, pk):
    if request.method == "POST":
        config = get_object_or_404(SamplingConfiguration, pk=pk)
        config.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"success": True})
        return redirect('config_preview')
    return JsonResponse({"success": False, "error": "Only POST allowed"})

def vendor_config(request):
    instrument_configs = InstrumentConfiguration.objects.all().order_by('instrument_num')
    return render(request, 'dashboard/config/vendor_config.html', {
        'instrument_configs': instrument_configs
    })


def vendor_config_create(request):
    if request.method == 'POST':
        instrument_name = request.POST.get('instrument_name', '').strip()
        instrument_num  = request.POST.get('instrument_num', '').strip()
        systerm_num     = request.POST.get('systerm_num', '').strip()
        upload_file     = request.FILES.get('upload_file')

        # ✅ 1) 去重校验：相同(仪器编号-仪器厂家-系统号)已存在则阻止创建
        exists = InstrumentConfiguration.objects.filter(
            instrument_name=instrument_name,
            instrument_num=instrument_num,
            systerm_num=systerm_num
        ).exists()

        if exists:
            # 复用已有的错误模板（dashboard/templates/dashboard/error.html）
            # 该模板已经在项目参数配置重复时使用过:contentReference[oaicite:2]{index=2}；模板在此:contentReference[oaicite:3]{index=3}
            return render(request, 'dashboard/error.html', {
                "message": f"已存在相同的配置：仪器编号 [{instrument_num}] / 厂家 [{instrument_name}] / 系统号 [{systerm_num}]，无需重复创建。"
            })

        # 通过校验后再创建
        instance = InstrumentConfiguration(
            instrument_name=instrument_name,
            instrument_num=instrument_num,
            systerm_num=systerm_num,
            upload_file=upload_file,
        )
        instance.save()
        return redirect('vendor_config')
    else:
        return render(request, 'dashboard/config/vendor_config_create.html')

    
def vendor_config_delete(request, pk):
    if request.method == "POST":
        config = get_object_or_404(InstrumentConfiguration, pk=pk)
        config.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"success": True})
        return redirect('vendor_config')
    return JsonResponse({"success": False, "error": "Only POST allowed"})

# 进样体积设置   
def injection_volume_config(request):
    injection_volume_configs = InjectionVolumeConfiguration.objects.all().order_by('project_name')
    return render(request, 'dashboard/config/injection_volume_config.html', {
        'injection_volume_configs': injection_volume_configs
    })

def injection_volume_config_create(request):
    if request.method == 'POST':
        instance = InjectionVolumeConfiguration(
            project_name=request.POST.get('project_name'),
            instrument_num=request.POST.get('instrument_num'),
            injection_volume=request.POST.get('injection_volume'),
            systerm_num=request.POST.get('systerm_num'),
        )

        instance.save()
        return redirect('injection_volume_config')
    else:
        return render(request, 'dashboard/config/injection_volume_config_create.html')

def injection_volume_config_delete(request, pk):
    if request.method == "POST":
        config = get_object_or_404(InjectionVolumeConfiguration, pk=pk)
        config.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"success": True})
        return redirect('config_preview')
    return JsonResponse({"success": False, "error": "Only POST allowed"})

# 进样盘号设置   
def injection_plate_config(request):
    configs = InjectionPlateConfiguration.objects.all().order_by('project_name')
    return render(request, 'dashboard/config/injection_plate_config.html', {
        'configs': configs
    })

# 进样盘号配置 —— 新建
def injection_plate_config_create(request):
    if request.method == 'POST':
        # 前端以 JSON 字符串传入 injection_plate_json，例如 ["Plate1","Plate2"]
        import json
        raw = request.POST.get('injection_plate_json', '[]')
        try:
            plate_list = json.loads(raw)
            # 简单清洗：去空、去重、全部转为字符串
            plate_list = list(dict.fromkeys([str(x).strip() for x in plate_list if str(x).strip()]))
        except Exception:
            plate_list = []

        instance = InjectionPlateConfiguration(
            project_name=request.POST.get('project_name', '').strip(),
            instrument_num=request.POST.get('instrument_num', '').strip(),
            systerm_num=request.POST.get('systerm_num'),
            injection_plate=plate_list,
        )
        instance.save()
        return redirect('injection_plate_config')
    else:
        return render(request, 'dashboard/config/injection_plate_config_create.html')


# 进样盘号配置 —— 删除
def injection_plate_config_delete(request, pk):
    if request.method == "POST":
        config = get_object_or_404(InjectionPlateConfiguration, pk=pk)
        config.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({"success": True})
        return redirect('injection_plate_config')
    return JsonResponse({"success": False, "error": "Only POST allowed"})


###################### Starlet专用函数 ###################### 
def _tail_number(s: str):
    """提取字符串末尾的连续数字；失败则返回 None"""
    import re
    if not s:
        return None
    m = re.search(r'(\d+)$', str(s).strip())
    return int(m.group(1)) if m else None


def _starlet_split_plates(scan_sheet, index_map):
    """
    将 Starlet 的扫码 sheet,按 TLabwareId 的末尾数字 分为多块板。
    返回 plate_groups = [(plate_no_int, row_indexes), ...]
    - plate_no_int: 1,2,3...
    - row_indexes: 属于该板的行下标列表（从 1 开始，不含表头）
    """
    labware_idx = index_map.get("TLabwareId")
    pos_idx     = index_map.get("TPositionId", 0)

    # 兜底：若缺 TLabwareId，退化为通过孔位序列回退分板（A1..H12 每 96 个一板）
    if labware_idx is None:
        nrows = scan_sheet.nrows
        all_rows = list(range(1, nrows))
        # 简单分块：每 96 行一板
        groups = [all_rows[i:i+96] for i in range(0, len(all_rows), 96)]
        return [(i+1, g) for i, g in enumerate(groups)]

    # 主规则：取 TLabwareId 的末尾数字做板号
    bucket = {}
    nrows  = scan_sheet.nrows
    for i in range(1, nrows):
        lab_str = str(scan_sheet.row_values(i)[labware_idx]).strip()
        k = _tail_number(lab_str)
        # 若无法提取，使用回退：按孔位序号（A1..H12）分组
        if k is None:
            # 根据 TPositionId 转序号（A1=1, A2=2, ... H12=96），超过 96 继续分块
            pos = str(scan_sheet.row_values(i)[pos_idx]).strip()
            if len(pos) >= 2 and pos[0].isalpha():
                row = "ABCDEFGH".find(pos[0].upper())
                try:
                    col = int(pos[1:])
                except:
                    col = 1
                if 0 <= row <= 7 and 1 <= col <= 12:
                    seq = row * 12 + col         # 1..96
                    k = ((seq-1) // 96) + 1      # 退化逻辑：一起视为第1板
                else:
                    k = 1
            else:
                k = 1
        bucket.setdefault(k, []).append(i)

    plate_groups = sorted(bucket.items(), key=lambda x: x[0])  # [(n, [rows...]), ...]
    return plate_groups


# 把原先 Starlet 单板的 96 孔对齐 与 定位孔补齐 逻辑封装进来
def _process_one_starlet_plate(scan_sheet, index_map, row_indexes, plate_no_int):
    """
    传入某一板对应的行下标 row_indexes，完成：
      1) 读出 Status/Barcode
      2) 96 孔对齐（A1..A12, B1..B12, ... H12）
      3) 定位孔补齐：第 1~12 板用 B1..B12；第 13~24 用 C1..C12；以此类推
    返回字典：
      {
        "Position": [...], "Status": [...], "OriginBarcode": [...],
        "CutBarcode": [...], "SubBarcode": [...], "Warm": [...],
        "plate_no_str": "X{n}"      # 导出/展示用
      }
    """
    # 列索引
    p_idx = index_map.get("TPositionId", 0)
    s_idx = index_map.get("TSumStateDescription", 0)
    b_idx = index_map.get("SPositionBC", 0)

    # ★ 新增：
    tsum_idx = index_map.get("TStatusSummary", None)
    tvol_idx = index_map.get("TVolume", None)

    # 先按 position -> row 信息字典
    row_by_pos = {}
    for i in row_indexes:
        pos = str(scan_sheet.row_values(i)[p_idx]).strip()
        row_by_pos[pos] = {
            "Status":        scan_sheet.row_values(i)[s_idx],
            "OriginBarcode": scan_sheet.row_values(i)[b_idx],
            "Warm":          "",   # Starlet 无 Warm

            # ★ 新增：把两列值也放到行记录里（可能为 None）
            "TStatusSummary": scan_sheet.row_values(i)[tsum_idx] if tsum_idx is not None else "",
            "TVolume":        scan_sheet.row_values(i)[tvol_idx] if tvol_idx is not None else "",
        }

    # 预期 96 孔顺序
    letters_fix = list("ABCDEFGH")
    nums_fix    = [str(i) for i in range(1, 13)]
    expected_positions = [f"{r}{c}" for r in letters_fix for c in nums_fix]

    # 对齐并在缺行时补齐
    Position, Status, OriginBarcode = [], [], []
    CutBarcode, SubBarcode, Warm    = [], [], []
    # ★ 新增：
    TStatusSummary, TVolume = [], []

    # 定位孔规则：第 1~12 板 -> B1..B12；13~24 -> C1..C12；……
    locator_rows = ["B","C","D","E","F","G","H"]  # 最多 7*12 = 84 块板可直接映射
    row_index = (plate_no_int - 1) // 12         # 第几行
    col_num   = ((plate_no_int - 1) % 12) + 1    # 第几列
    locator_target = None
    if row_index < len(locator_rows):
        locator_target = f"{locator_rows[row_index]}{col_num}"  # 如 B1, B2, ... C1, ...

    for pos in expected_positions:
        if pos in row_by_pos:
            status = row_by_pos[pos]["Status"]
            bc     = row_by_pos[pos]["OriginBarcode"]
            warm   = row_by_pos[pos]["Warm"]
            # ★ 新增：
            tsum   = row_by_pos[pos]["TStatusSummary"]
            tvol   = row_by_pos[pos]["TVolume"]
        else:
            status, bc, warm = "Not used", "NOTUBE", ""
            # ★ 新增：缺行时默认
            tsum, tvol = "", 0

            # 缺行但命中定位孔位置时，用 X{n}
            if locator_target and pos == locator_target:
                bc = f"X{plate_no_int}"

        Position.append(pos)
        Status.append(status)
        Warm.append(warm)
        # ★ 新增：
        TStatusSummary.append(tsum)
        TVolume.append(tvol)

        bc_str = str(bc)
        parts  = bc_str.split("-", 1)
        OriginBarcode.append(bc_str)
        CutBarcode.append(parts[0])
        SubBarcode.append("-" + parts[1] if len(parts) == 2 else "")

    return {
        "Position": Position,
        "Status": Status,
        "OriginBarcode": OriginBarcode,
        "CutBarcode": CutBarcode,
        "SubBarcode": SubBarcode,
        "Warm": Warm,
        "plate_no_str": f"X{plate_no_int}",

        # ★ 新增：
        "TStatusSummary": TStatusSummary,
        "TVolume": TVolume,
    }


def format_vialpos_column(df: pd.DataFrame, colname: str = "VialPos"):
    """
    统一格式化 VialPos 列：
    - 空值 -> ""
    - 1.0 -> "1"
    - 86.0 -> "86"
    - 3 -> "3"
    """
    def fmt(v):
        if pd.isna(v):
            return ""                     # 也可以 return "nan" 看你需求
        try:
            # 能转成数字的全部按整数格式输出
            return str(int(float(v)))
        except Exception:
            # 极少数无法解析成数字的，保持原样
            return str(v)

    df[colname] = df[colname].map(fmt)
    return df

# 结果处理，用户在前端功能入口处选择项目，上传文件并点击提交按钮后的处理逻辑 locator
def ProcessResult(request):
    """
    同时支持：
      - NIMBUS：单板
      - Starlet：单板/多板（按 TLabwareId 末尾数字分组）
    """
    from collections import defaultdict, Counter, deque
    import os, re
    import xlrd
    import pandas as pd
    from io import StringIO
    from django.utils import timezone
    from django.http import HttpResponse, HttpResponseBadRequest
    from django.shortcuts import get_object_or_404, render

    # ========== 1. 入参与上传文件 ==========
    project_id      = request.POST.get("project_id")
    project_name = request.POST.get("project_name")
    platform        = request.POST.get("platform")                # 'NIMBUS' | 'Starlet'
    injection_plate = request.POST.get("injection_plate") if 'injection_plate' in request.POST else None
    instrument_num  = request.POST.get("instrument_num")  # 默认上机仪器
    systerm_num  = request.POST.get("systerm_num")  # 系统号
    testing_day      = request.POST.get("testing_day") # 检测日期

    if testing_day == "today":
        today_str  = timezone.localtime().strftime("%Y%m%d")  # 用于96孔板和上机列表
        record_date = date.today()  # 用于历史标本查找和统计
    else:
        today_str = (timezone.localtime() + timedelta(days=1)).strftime("%Y%m%d")
        record_date = date.today() + timedelta(days=1)

    Stationlist = request.FILES.get('station_list')               # 每日操作清单
    Scanresult  = request.FILES.get('scan_result')                # 扫码结果
    if not (Stationlist and Scanresult and project_id and platform and instrument_num):
        return HttpResponseBadRequest("缺少必要参数或文件。")

    Stationtable = xlrd.open_workbook(filename=None, file_contents=Stationlist.read())
    Scantable    = xlrd.open_workbook(filename=None, file_contents=Scanresult.read())

    scan_sheet   = Scantable.sheets()[0]

    # 扫码表头索引
    nrows = scan_sheet.nrows
    ncols = scan_sheet.ncols
    scan_header = [str(scan_sheet.row_values(0)[i]).strip() for i in range(ncols)]
    scan_index  = {col: idx for idx, col in enumerate(scan_header)}

    # 关键列索引（容错：缺失时给默认 0）
    POS_IDX   = scan_index.get("TPositionId", 0)
    STAT_IDX  = scan_index.get("TSumStateDescription", 0)
    BC_IDX    = scan_index.get("SPositionBC", 0)
    WARM_IDX  = scan_index.get("Warm")               # NIMBUS有；Starlet通常无
    LABW_IDX  = scan_index.get("TLabwareId")        # Starlet有

    # ========== 2. 基础配置与映射（整次仅读一次） ==========

    # 获取后台设置的项目参数，如果没设置报错并提示
    try:
        config = SamplingConfiguration.objects.get(project_name=project_name,default_upload_instrument=instrument_num,systerm_num=systerm_num)
    except SamplingConfiguration.DoesNotExist:
        # 返回友好的错误提示页面
        return render(request, "dashboard/error.html", {
            "message": "未配置项目参数，请前往后台参数配置中完善该项目设置后重试。"
        })

    project_name   = config.project_name
    project_name_full   = config.project_name_full
    mapping_path   = config.mapping_file.path
    df_mapping_wc  = pd.read_excel(mapping_path, sheet_name="工作清单")   # for worksheet
    df_worklistmap = pd.read_excel(mapping_path, sheet_name="上机列表")    # worklist mapping 模板

    # 解析后台设置的上机模板（txt/csv）→ DataFrame（只需列名 / txt_headers）,获取表头
    try:
        instrument_config = InstrumentConfiguration.objects.get(instrument_num=instrument_num,systerm_num=systerm_num)
    except InstrumentConfiguration.DoesNotExist:
        # 返回友好的错误提示页面
        return render(request, "dashboard/error.html", {
            "message": "未配置仪器厂家参数，请前往后台参数配置中完善设置后重试。"
        })

    instrument_name = instrument_config.instrument_name
    if not instrument_config.upload_file:
        return HttpResponse("未设置该上机仪器对应的上机模板,请设置后再试", status=404)

    ext = os.path.splitext(instrument_config.upload_file.name)[1].lower()
    if ext not in (".txt", ".csv"):
        return HttpResponse(f"仅支持 .txt 或 .csv 模板，当前为：{ext}", status=400)

    with instrument_config.upload_file.open("rb") as f:
        raw = f.read()
    text, last_err = None, None
    for enc in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError as e:
            last_err = e
    if text is None:
        raise last_err or UnicodeDecodeError("unknown", b"", 0, 1, "unable to decode upload_file")
    df_template = pd.read_csv(StringIO(text), sep=None, engine="python", dtype=str)
    txt_headers = df_template.columns.tolist()

    # 岗位清单主条码 ↔ 实验号（获取一一对应关系）
    st_sheet   = Stationtable.sheets()[0]
    st_nrows   = st_sheet.nrows
    st_ncols   = st_sheet.ncols
    st_header  = [str(st_sheet.row_values(0)[i]).strip() for i in range(st_ncols)]
    st_index   = {col: idx for idx, col in enumerate(st_header)}
    MB_IDX     = st_index.get("主条码", 0)
    SN_IDX     = st_index.get("实验号", 0)
    MainBarcode = [str(st_sheet.row_values(i)[MB_IDX]) for i in range(1, st_nrows)]
    SampleName  = [str(st_sheet.row_values(i)[SN_IDX]) for i in range(1, st_nrows)]
    barcode_to_names = defaultdict(list)
    for bc, sn in zip(MainBarcode, SampleName):
        barcode_to_names[str(bc)].append(str(sn))


    # 保存岗位清单中主条码和实验号映射的json数据（每天一份，不能重复）
    def _station_store_path(day_date: date) -> str:
        day_dir = os.path.join(settings.DOWNLOAD_ROOT, "岗位清单", day_date.strftime("%Y-%m-%d"))
        os.makedirs(day_dir, exist_ok=True)
        return os.path.join(day_dir, "station_list.json")

    def _load_station_store(path: str) -> dict:
        if not os.path.isfile(path):
            # 保证 warning 在最前
            return {"warning": [], "主条码->实验号列表": {}, "实验号->主条码": {}}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            # 兼容缺字段
            if "warning" not in data:
                data = {"warning": [], **data}
            data.setdefault("主条码->实验号列表", {})
            data.setdefault("实验号->主条码", {})
            return data
        except Exception:
            # 文件坏了就重新来（不阻断主流程）
            return {"warning": [{"type": "load_error", "msg": f"读取失败，已重置：{os.path.basename(path)}"}],
                    "主条码->实验号列表": {}, "实验号->主条码": {}}
    
    def _dedup_warning(existing: list, new_items: list) -> list:
        # 用 (type, exp, old, new) 作为去重 key，避免 warning 越堆越多
        seen = set()
        out = []
        for w in (existing or []) + (new_items or []):
            if not isinstance(w, dict):
                continue
            key = (w.get("type"), w.get("experiment_no"), w.get("old_main_barcode"), w.get("new_main_barcode"), w.get("msg"))
            if key in seen:
                continue
            seen.add(key)
            out.append(w)
        return out


    def _merge_station_maps(store: dict, new_map: dict) -> tuple[dict, dict]:
        """
        store:
          - 主条码->实验号列表: dict[str, list[str]]
          - 实验号->主条码: dict[str, str]
          - warning: list[dict]
        new_map: barcode_to_names (dict[str, list[str]])
        返回：(new_store, summary)
        """
        mb2sn = store.get("主条码->实验号列表", {}) or {}
        sn2mb = store.get("实验号->主条码", {}) or {}
        warnings_new = []

        added_pairs = 0
        conflict_cnt = 0

        for mb, sns in (new_map or {}).items():
            mb = (mb or "").strip()
            if not mb:
                continue

            # 规范化 sns：去空/去重/保持顺序
            seen_sn = set()
            norm_sns = []
            for sn in (sns or []):
                sn = (str(sn) or "").strip()
                if not sn or sn in seen_sn:
                    continue
                seen_sn.add(sn)
                norm_sns.append(sn)
            if not norm_sns:
                continue

            exist_list = mb2sn.get(mb, [])
            if not isinstance(exist_list, list):
                exist_list = [str(exist_list)]
            exist_set = set(str(x).strip() for x in exist_list if str(x).strip())

            # 逐实验号合并，并做“实验号->主条码”冲突检测
            for sn in norm_sns:
                if sn in sn2mb:
                    # 已存在映射：若冲突则记录 warning 并跳过
                    old_mb = str(sn2mb.get(sn) or "").strip()
                    if old_mb and old_mb != mb:
                        conflict_cnt += 1
                        warnings_new.append({
                            "type": "conflict",
                            "experiment_no": sn,
                            "old_main_barcode": old_mb,
                            "new_main_barcode": mb,
                            "msg": f"实验号 {sn} 当天已映射主条码 {old_mb}，本次上传为 {mb}，已忽略本次冲突行"
                        })
                        continue
                    # old_mb == mb：完全重复 -> 忽略
                else:
                    # 新实验号：写入 sn->mb
                    sn2mb[sn] = mb

                # mb -> sn list：不存在则追加
                if sn not in exist_set:
                    exist_list.append(sn)
                    exist_set.add(sn)
                    added_pairs += 1

            mb2sn[mb] = exist_list
                    
    
        # warning 合并且去重（warning 必须在 json 最前面）
        store["warning"] = _dedup_warning(store.get("warning", []), warnings_new)
        store["主条码->实验号列表"] = mb2sn
        store["实验号->主条码"] = sn2mb

        summary = {
            "added_pairs": added_pairs,
            "conflicts": conflict_cnt,
            "saved": False,
        }
        return store, summary


    station_save_summary = None
    generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
    try:
        store_path = _station_store_path(record_date)  # record_date 你前面已计算（today/tomorrow）
        store = _load_station_store(store_path)
        store2, summary = _merge_station_maps(store, barcode_to_names)

        # 只有真的有新增/新增 warning 才写盘（避免无意义改动）
        need_write = (summary["added_pairs"] > 0) or (summary["conflicts"] > 0) or (not os.path.isfile(store_path))
        if need_write:
            ordered_out = {
                "生成时间": generated_at,
                "warning": store2.get("warning", []),
                "主条码->实验号列表": store2.get("主条码->实验号列表", {}),
                "实验号->主条码": store2.get("实验号->主条码", {}),
            }
            with open(store_path, "w", encoding="utf-8") as f:
                json.dump(ordered_out, f, ensure_ascii=False, indent=2)

            summary["saved"] = True

        station_save_summary = {
            **summary,
            "path": store_path,
        }
        logging.getLogger(__name__).warning(
            "[station_list saved] path=%s added_pairs=%s conflicts=%s",
            store_path, summary["added_pairs"], summary["conflicts"]
        )
    except Exception as e:
        # 任何保存异常：不阻断主流程
        logging.getLogger(__name__).warning("[station_list save failed] %s", e)
        station_save_summary = {"saved": False, "added_pairs": 0, "conflicts": 0, "error": str(e)}


    # 曲线/质控映射（获取一一对应关系,供后续识别非临床样本,即曲线和质控）
    barcode_to_name = dict(zip(df_mapping_wc["Barcode"].astype(str), df_mapping_wc["Name"]))

    # 用于构建曲线/QC/Test/DB 序列
    test_count    = config.test_count
    curve_points  = config.curve_points
    df_std = df_mapping_wc.copy()
    df_std["Code"] = df_std["Code"].astype(str)
    df_std = df_std[df_std["Code"].str.match(r"^STD\d+$", na=False)].copy()
    df_std["__std_idx"] = df_std["Code"].str.replace("STD", "", regex=False).astype(int)
    df_std = df_std.sort_values("__std_idx")
    std_names = df_std["Name"].head(curve_points + 1).tolist()
    if not std_names or len(std_names) < (curve_points + 1):
        std_names = [f"STD{i}" for i in range(curve_points + 1)]

    qc_names = df_mapping_wc.loc[df_mapping_wc["Code"].astype(str).str.startswith("QC"), "Name"].unique().tolist()

    # 进样体积（非必须设置项）
    try:
        injection_cfg = InjectionVolumeConfiguration.objects.get(project_name=project_name,instrument_num=instrument_num,systerm_num=systerm_num)
        injection_vol  = injection_cfg.injection_volume
    except InjectionVolumeConfiguration.DoesNotExist:
        injection_vol  = ""


    # ====== 工具：把“对齐后的单板数据” → 构建（工作清单 + 上机列表）并返回一个 plate payload ======
    # 适用于只有一块板的情况，多块板需要循环反复调用此函数
    def build_one_plate_payload(aligned: dict, layout: str, plate_no_str: str):
        """
        aligned: 由 _process_one_starlet_plate()（Starlet）或 NIMBUS 构造出的：
          Position/Status/OriginBarcode/CutBarcode/SubBarcode/Warm
        layout: 'horizontal' | 'vertical'
        plate_no_str: "X{n}"
        """
        Position       = aligned["Position"]
        Status         = aligned["Status"]
        OriginBarcode  = aligned["OriginBarcode"]
        CutBarcode     = aligned["CutBarcode"]
        SubBarcode     = aligned["SubBarcode"]
        Warm           = aligned["Warm"]

        # ★ 新增：
        TStatusSummary = aligned.get("TStatusSummary", [""] * len(Position))
        TVolume        = aligned.get("TVolume",        [0]   * len(Position))

        # 96 补齐（Starlet单块板不到96个孔位，需补齐以保证数据格式与NIMBUS统一）
        rows, cols = 8, 12
        TOTAL = rows * cols
        def pad_to(seq, n, fill=""):
            if len(seq) < n: seq.extend([fill] * (n - len(seq)))
        pad_to(Position, TOTAL, "")
        pad_to(OriginBarcode, TOTAL, "NOTUBE")
        pad_to(CutBarcode, TOTAL, "")
        pad_to(SubBarcode, TOTAL, "")
        pad_to(Warm, TOTAL, "")
        pad_to(Status, TOTAL, "Not used")

        # ★ 新增：
        pad_to(TStatusSummary, TOTAL, 0)
        pad_to(TVolume,        TOTAL, 0) 

        # —— 与岗位清单匹配（按原来的规则）——
        cut_counter = Counter(CutBarcode)
        MatchSampleName, MatchResult = [], []
        DupBarcode, DupBarcodeSampleName = [], []
        for cb in CutBarcode:
            cb_str = str(cb)
            matched_names = barcode_to_names.get(cb_str, [])
            cb_count = cut_counter[cb_str]
            if len(matched_names) == 1:
                MatchResult.append("TRUE")
                MatchSampleName.append(matched_names[0]); DupBarcode.append(""); DupBarcodeSampleName.append("")
            elif len(matched_names) == 0:
                MatchResult.append("FALSE")
                if cb_str != "":
                    MatchSampleName.append(cb_str); DupBarcode.append(""); DupBarcodeSampleName.append("")
                else:
                    MatchSampleName.append("");    DupBarcode.append(""); DupBarcodeSampleName.append("")
            elif len(matched_names) == 2:
                MatchResult.append("TRUE")
                if matched_names[0] == matched_names[1]:
                    DupBarcode.append("Likely"); MatchSampleName.append(matched_names[0]); DupBarcodeSampleName.append("")
                else:
                    DupBarcode.append("TRUE");   MatchSampleName.append(matched_names[0] + "-" + matched_names[1])
                    DupBarcodeSampleName.append("TRUE" if cb_count >= 2 else "")
            else:
                MatchResult.append("TRUE")
                unique_Middlelist = list(dict.fromkeys(matched_names))
                order = {'VF': 1, 'AE': 2, 'VD': 3, 'V': 4, 'VK': 5, 'WV': 6}
                def sort_key(x):
                    for prefix_length in (2, 1):
                        prefix = x[:prefix_length]
                        if prefix in order: return order[prefix]
                    return len(order) + 1
                sorted_lis = sorted(unique_Middlelist, key=sort_key)
                if len(unique_Middlelist) >= 2:
                    MatchSampleName.append('-'.join(sorted_lis)); DupBarcode.append("TRUE")
                    DupBarcodeSampleName.append("TRUE" if cb_count >= 2 else "")
                else:
                    MatchSampleName.append(matched_names[0]); DupBarcode.append(""); DupBarcodeSampleName.append("")

        # 二 构建 96 孔工作清单网格（保持原渲染结构）——
        letters = list("ABCDEFGH"); nums = [str(i) for i in range(1, 13)]
        # 定位孔集合（NIMBUS: Warm 含 X；Starlet: _process_one_starlet_plate 已把 Xn 放在 OriginBarcode 且 Warm 为空）
        locator_positions = set()
        if platform == "NIMBUS":
            for idx, w in enumerate(Warm):
                if isinstance(w, str) and "X" in w:
                    locator_positions.add(Position[idx])

        barcode_to_well = defaultdict(deque)  # 把普通字典改成 defaultdict(deque)，用于保存 “一个条码 → 多个 (孔位,孔号)”

        # barcode_to_well = {}   # OriginBarcode -> (well_str, well_number)
        worksheet_grid  = [[None for _ in nums] for _ in letters]
        error_rows      = []

        def _well_number_rowwise(row_idx: int, col: int) -> int:
            return row_idx * 12 + col

        def build_well(row_letter, col_num, row_idx, col_idx, data_idx, well_index):
            well_pos_str = f"{row_letter}{col_num}"
            origin_barcode = str(OriginBarcode[data_idx])

            if origin_barcode not in ("", "nan"):
                # 同一条码可能对应多个孔位，全部记录到队列里
                barcode_to_well[origin_barcode].append((well_pos_str, well_index))

            value = str(MatchSampleName[data_idx])
            match_sample = value if MatchResult[data_idx] == "TRUE" else ("" if value == "" else barcode_to_name.get(value, "No match"))

            is_locator    = (well_pos_str in locator_positions) or (str(OriginBarcode[data_idx]).upper().startswith("X"))
            locator_label = Warm[data_idx] if is_locator else ""
            if is_locator and (not locator_label):
                locator_label = plate_no_str

            well = {
                "letter": row_letter, 
                "num": col_num, 
                "well_str": well_pos_str, 
                "index": well_index,
                
                "locator": is_locator, 
                "locator_warm": locator_label,
                "match_sample": match_sample,
                "cut_barcode": CutBarcode[data_idx], "sub_barcode": SubBarcode[data_idx],
                "origin_barcode": OriginBarcode[data_idx],
                "warm": Warm[data_idx], "status": Status[data_idx],
                "dup_barcode": DupBarcode[data_idx], "dup_barcode_sample": DupBarcodeSampleName[data_idx],
            }

            status_text = str(Status[data_idx]).strip().lower()

            # === Starlet 新口径：用 TStatusSummary + TVolume 判断两类报错 ===
            def _to_int(x, default=0):
                try:
                    # 兼容 float/str/'16384.0' 等
                    return int(float(x))
                except Exception:
                    return default

            tsum = _to_int(TStatusSummary[data_idx], 0)
            tvol = _to_int(TVolume[data_idx], 0)

            # 记录SampleRecord模型中的error_info字段
            error_info = ""

            # 1) 吸液报错（孔位可用）：tsum=16384 且 tvol=0
            if tsum == 16384 and tvol == 0:
                error_info = "1"
                well["flag_suck"] = "1"
                error_rows.append({
                    "sample_name":   match_sample,
                    "origin_barcode": OriginBarcode[data_idx],
                    "plate_no":       plate_no_str,
                    "well_str":       well_pos_str,
                    "warn_level":     "1",
                    "warn_info":      "Not used",
                })

            # 2) 打液报错（孔位不可用）：tsum=16384 且 tvol!=0
            elif tsum == 16384 and tvol != 0:
                error_info = "16384"
                well["flag_dispense"] = "1"
                error_rows.append({
                    "sample_name":   match_sample,
                    "origin_barcode": OriginBarcode[data_idx],
                    "plate_no":       plate_no_str,
                    "well_str":       well_pos_str,
                    "warn_level":     "16384",
                    "warn_info":      "Pipetting error",
                })

            row_data = {
                "sample_name": match_sample,
                "origin_barcode": OriginBarcode[data_idx],
                "plate_no": plate_no_str,
                "well_str": well_pos_str,
                "warn_level": Warm[data_idx],
                "warn_info": Status[data_idx],
            }

            if (str(Warm[data_idx]) in ["1", "4", "16384"]) or (well_pos_str == "H12") or (match_sample == "No match"):
                # ✅ 如果是 H12，放到第一行
                if well_pos_str == "H12":
                    error_rows.insert(0, row_data)
                else:
                    error_rows.append(row_data)

            # 落库（以“同日-同项目-同板号-孔位”为粒度）
            SampleRecord.objects.update_or_create(
                project_name=project_name,
                record_date=record_date,
                plate_no=plate_no_str,
                well_str=well_pos_str,
                defaults={
                    "sample_name": match_sample,
                    "barcode": origin_barcode,
                    "error_info": error_info,
                }
            )
            return well

        # 清理同日同项目同板号旧记录（避免重复）
        SampleRecord.objects.filter(project_name=project_name, record_date=record_date, plate_no=plate_no_str).delete()

        if layout == 'vertical':   # NIMBUS（列优先）
            for col_idx, col_num in enumerate(nums):
                for row_idx, row_letter in enumerate(letters):
                    data_idx   = col_idx * 8 + row_idx
                    well_index = _well_number_rowwise(row_idx, int(col_num))
                    worksheet_grid[row_idx][col_idx] = build_well(row_letter, col_num, row_idx, col_idx, data_idx, well_index)

        else:                      # Starlet（行优先）
            for row_idx, row_letter in enumerate(letters):
                for col_idx, col_num in enumerate(nums):
                    data_idx   = row_idx * 12 + col_idx
                    well_index = _well_number_rowwise(row_idx, int(col_num))
                    worksheet_grid[row_idx][col_idx] = build_well(row_letter, col_num, row_idx, col_idx, data_idx, well_index)

        worksheet_table = [[worksheet_grid[r][c] for c in range(12)] for r in range(8)]

        # —— 生成上机列表 —— #
        # ClinicalSample：OriginBarcode 中不在映射表“Barcode”里的条码（若 Warm 含 X 则以 Xn 记名）
        mapping_barcodes = set(str(x) for x in df_mapping_wc["Barcode"].tolist())
        ClinicalSample = []
        for i, ob in enumerate(OriginBarcode):
            ob_str   = str(ob)
            warm_val = str(Warm[i]).strip().upper()
            if ob_str not in mapping_barcodes:
                if 'X' in warm_val: ClinicalSample.append(warm_val)  # 定位孔
                else:               ClinicalSample.append(ob_str)

        test_list   = ["DB1"] + [f"Test{i}" for i in range(test_count)]
        curve_list  = ["DB2"] + std_names
        qc_list1    = ["DB3"] + qc_names + ["DB4"]
        qc_list2    = qc_names + ["DB5"]
        SampleName_list = test_list + curve_list + qc_list1 + ClinicalSample + qc_list2
        SampleName_list = [name for name in SampleName_list if isinstance(name, str) and name.count('-') <= 3]

        # worklist 空表
        worklist_table = pd.DataFrame(columns=df_template.columns)
        worklist_table[worklist_table.columns[0]] = SampleName_list

        # value 队列：Name -> [barcodes...]
        name_to_barcodes = defaultdict(deque)
        for barcode, name in barcode_to_name.items():
            name_to_barcodes[name].append(barcode)

        first_col = worklist_table.columns[0]

        # 记录哪些列需要镜像第一列（由映射表的 * 标注）
        mirror_cols = set()

        for _, row in df_worklistmap.iterrows():
            sample_key = row.iloc[0]
            fill_vals  = row.iloc[1:]

            # ===== 新增：标记本行是否为“默认 * 行” =====
            is_default_star_row = (str(sample_key).strip() == "*")

            def fill_cols(mask):
                for col, val in zip(worklist_table.columns[1:], fill_vals.values):
                    # ------- 新增：当“当前行是 * 行 且 该列映射值也为 *” → 镜像第一列 -------
                    # ★ 标注“需要镜像”的列（映射表该列写了 "*"）
                    if str(val).strip() == "*":
                        mirror_cols.add(col)
                        continue
                    # -------------------------------------------------------------------

                    if col in ("SmplInjVol", "Injection volume"):
                        continue
                    if col in ("VialPos", "Vial position", "样品瓶"):
                        ROWS = list("ABCDEFGH")
                        def _resolve_vialpos(sample_name_value):
                            s = str(sample_name_value).strip().upper()
                            m = re.fullmatch(r"X(\d+)", s)
                            if m:
                                k0 = int(m.group(1)) - 1
                                if platform == "NIMBUS":
                                    coln     = 3 + (k0 // 8)
                                    row_idx  = k0 % 8
                                else:  # Starlet
                                    row_idx  = 1 + (k0 // 12)
                                    coln     = 1 + (k0 % 12)
                                if not (0 <= row_idx < 8 and 1 <= coln <= 12):
                                    return None
                                well_pos = f"{ROWS[row_idx]}{coln}"
                                well_no  = _well_number_rowwise(row_idx, coln)
                                if instrument_name == "Thermo" or instrument_name == "Agilent":
                                    if val == "{{Well_Number}}":
                                        return f"{injection_plate}:{well_no}" if injection_plate else well_no
                                    else:
                                        return f"{injection_plate}-{well_pos}" if injection_plate else well_pos
                                else:
                                    if val == "{{Well_Number}}":
                                        return well_no
                                    else:
                                        return ell_pos

                            if val in ["{{Well_Number}}", "{{Well_Position}}"]:
                                # 1) QC/STD：通过 name->barcode 队列取条码，再查位置信息
                                if sample_name_value in name_to_barcodes and name_to_barcodes[sample_name_value]:
                                    barcode = name_to_barcodes[sample_name_value].popleft()
                                    wells_q = barcode_to_well.get(barcode)
                                    if wells_q:
                                        # 依次使用该条码对应的各个孔位
                                        pos, no = wells_q.popleft()
                                        # 下面保持你原来的仪器判断逻辑不变
                                        if instrument_name == "Thermo" or instrument_name == "Agilent":
                                            if val == "{{Well_Number}}":
                                                return f"{injection_plate}:{no}" if injection_plate else no
                                            else:
                                                return f"{injection_plate}-{pos}" if injection_plate else pos
                                        else:
                                            if val == "{{Well_Number}}":
                                                return no
                                            else:
                                                return pos

                                # 2) 临床样本：第一列就是条码
                                elif sample_name_value in barcode_to_well:
                                    wells_q = barcode_to_well.get(str(sample_name_value))
                                    if wells_q:
                                        pos, no = wells_q.popleft()
                                        if instrument_name == "Thermo" or instrument_name == "Agilent":
                                            if val == "{{Well_Number}}":
                                                return f"{injection_plate}:{no}" if injection_plate else no
                                            else:
                                                return f"{injection_plate}-{pos}" if injection_plate else pos
                                        else:
                                            if val == "{{Well_Number}}":
                                                return no
                                            else:
                                                return pos

                                else:              
                                    if instrument_name == "Thermo" or instrument_name == "Agilent":
                                        if val == "{{Well_Number}}":
                                            return f"{injection_plate}:{'1'}" if injection_plate else "1"
                                        else:
                                            return f"{injection_plate}-{'A1'}" if injection_plate else "A1"
                                    else:
                                        if val == "{{Well_Number}}":
                                            return "1"
                                        else:
                                            return "A1"

                            return val

                        worklist_table.loc[mask, col] = worklist_table.loc[mask, first_col].apply(_resolve_vialpos)

                    else:
                        worklist_table.loc[mask, col] = val

            if str(sample_key) == "DB*":
                mask = worklist_table.iloc[:, 0].str.startswith("DB")
                fill_cols(mask)
            elif str(sample_key).startswith("DB"):
                mask = worklist_table.iloc[:, 0] == str(sample_key)
                fill_cols(mask)

            elif str(sample_key) == "Test*":
                mask = worklist_table.iloc[:, 0].str.startswith("Test")
                fill_cols(mask)
            elif str(sample_key).startswith("Test"):
                mask = worklist_table.iloc[:, 0] == str(sample_key)
                fill_cols(mask)


            elif str(sample_key) == "STD*":
                mask = worklist_table.iloc[:, 0].str.startswith("STD")
                fill_cols(mask)
            elif str(sample_key).startswith("STD"):
                mask = worklist_table.iloc[:, 0] == str(sample_key)
                fill_cols(mask)

            elif str(sample_key) == "*":
                mask = worklist_table.iloc[:, 1].isna()
                fill_cols(mask)

        year       = today_str[:4]
        yearmonth  = today_str[:6]

        setname    = f"{instrument_num}_{systerm_num}_{project_name}_{today_str}_{plate_no_str}_GZ"
        output_val = f"{year}\\{yearmonth}\\Data{setname}"

        if "SetName" in worklist_table.columns:  worklist_table["SetName"]  = setname
        if "OutputFile" in worklist_table.columns: worklist_table["OutputFile"] = output_val

        # Thermo 专用：若第一列存在完全相同的多行值，则改成 原值-1、原值-2、…（仅对重复值生效）——
        # 判断当前仪器是否 Thermo（InstrumentConfiguration.instrument_name 中包含 "Thermo"）
        vendor_name = str(getattr(instrument_config, "instrument_name", "")).lower()
        if "thermo" in vendor_name:
            # 第一列的列名
            first_col_name = worklist_table.columns[0]
            # 统一按字符串处理
            s = worklist_table[first_col_name].astype(str)

            # 统计每个值出现次数，用于只对“重复值”做处理
            vc = s.map(s.value_counts())
            # 对同值分组做累加计数：0,1,2,...  -> 我们需要 1,2,3,...
            order = s.groupby(s).cumcount() + 1

            # 只改“出现次数 > 1”的那些值；出现 1 次的保持不变
            mask = vc > 1
            worklist_table.loc[mask, first_col_name] = s[mask] + "-" + order[mask].astype(str)

        # ★ 统一“镜像列”赋值（整列镜像，不再区分哪些行）
        if mirror_cols:
            first_col_name = worklist_table.columns[0]
            for col in mirror_cols:
                worklist_table[col] = worklist_table[first_col_name]

        # 进样体积
        for col in ["SmplInjVol", "Injection volume"]:
            if col in worklist_table.columns:
                worklist_table[col] = injection_vol
        
        # 进样盘
        if "PlatePos" in worklist_table.columns:
            worklist_table["PlatePos"] = injection_plate

        # 将第一列内容中含有‘X’关键词的行（即定位孔），整行移到第一列内容为‘DB4’的这一行后
        # X 行
        col = worklist_table.columns[0]
        x_rows = worklist_table[worklist_table[col].astype(str).str.contains('X', na=False)]

        # DB4 行索引
        db4_idx = worklist_table.index[worklist_table[col] == 'DB4'][0]

        # 原表删除 X 行
        df = worklist_table.drop(x_rows.index)

        # 重新拼接（看起来像“原地替换”）
        worklist_table = pd.concat(
            [df.loc[:db4_idx], x_rows, df.loc[db4_idx + 1:]],
            ignore_index=True
        )

        # 过滤掉 报错信息表 warn_level == 16384 的条码
        exclude_barcodes = {
            str(r.get("origin_barcode")).strip()
            for r in error_rows
            if str(r.get("warn_level")) == "16384"
        }

        if exclude_barcodes:
            first_col_name = worklist_table.columns[0]  # 上机列表第一列
            # 统一按字符串比对；将第一列等于这些条码的行删除
            mask_keep = ~worklist_table[first_col_name].astype(str).isin(exclude_barcodes)
            worklist_table = worklist_table[mask_keep]

        # IGF1项目需要删除以 DB 和 Test 开头的行，但保留 DB3。同时将 DB3 替换为 Blank，然后把STD0的孔位赋值给Blank
        if project_name == 'IGF1':
            # 获取第一列列名（不管叫什么）
            col = worklist_table.columns[0]
            col2 = "Sample name"
            col3 = "Vial position"

            val = worklist_table.loc[worklist_table[col] == "STD0", col3].iat[0]

            # 1. 删除以 DB 和 Test 开头的行，但保留 DB3
            worklist_table = worklist_table[~(
                (worklist_table[col].str.startswith("DB", na=False) & (worklist_table[col] != "DB3")) |
                (worklist_table[col].str.startswith("Test", na=False))
            )]
            
            # 2. 将 DB3 替换为 Blank
            worklist_table.loc[worklist_table[col] == "DB3", col] = "Blank"
            worklist_table.loc[worklist_table[col2] == "DB3", col2] = "Blank"

            # 把STD0的孔位赋值给Blank
            worklist_table.loc[worklist_table[col] == "Blank", col3] = val

        # AE项目上机时，需将临床样品的条码替换为实验号  
        is_ae_project = ("AE" in str(project_name).upper())

        MAIN_BARCODE_RE = re.compile(r"^(\d{8,})")  # 8位以上数字开头，按你条码长度可调

        def _main_barcode(val: str) -> str:
            """
            从 '01000086089-01' 提取主条码 '01000086089'
            若不是数字条码，返回空串
            """
            s = str(val).strip()
            m = MAIN_BARCODE_RE.match(s)
            return m.group(1) if m else ""

        # if is_ae_project:
        #     first_col = worklist_table.columns[0]

        #     def _to_exp_no(val):
        #         raw = str(val).strip()
        #         if not raw:
        #             return raw

        #         mb = _main_barcode(raw)
        #         if not mb:
        #             # X1 / DB1 / STD0 等
        #             return raw

        #         names = barcode_to_names.get(mb)
        #         if names:
        #             # ✅ 只替换显示，保留子条码后缀（避免用户看不出是哪支）
        #             # 例如：VD5674（01000086089-01）
        #             return f"{names[0]}"

        #         return raw

        #     worklist_table[first_col] = worklist_table[first_col].map(_to_exp_no)

        if 'VialPos' in worklist_table.columns:
            worklist_table = format_vialpos_column(worklist_table, "VialPos")
        
        if '级别' in worklist_table.columns:
            worklist_table = format_vialpos_column(worklist_table, "级别")

        worklist_table = worklist_table.fillna("")
        worklist_records = worklist_table.to_dict(orient="records")

        header_meta = {
            "test_date": timezone.localtime().strftime("%Y-%m-%d"),
            "plate_no": plate_no_str,
            "instrument_num": instrument_num,
            "systerm_num": systerm_num,
            "injection_plate": injection_plate,
            "today_str": today_str,
        }

        return {
            "project_name": project_name,
            "project_name_full": project_name_full,
            "instrument_num": instrument_num,
            "systerm_num": systerm_num,
            "plate_no": plate_no_str,
            "platform": platform,
            "today_str": today_str,
            "testing_day": testing_day,
           
            "worksheet_table": worksheet_table,
            "error_rows": error_rows,
            "txt_headers": txt_headers,
            "worklist_records": worklist_records,
            "header": header_meta,

        }

    # ========== 3. 分板 & 逐板处理 ==========
    plates_payload = []

    if platform == "NIMBUS":
        # —— NIMBUS：仍按“单板”处理 —— #
        # 直接从整表读取出 arrays（行顺序不调，后续在 build_one_plate_payload 里补齐&渲染）
        Position, Status, OriginBarcode, CutBarcode, SubBarcode, Warm = [], [], [], [], [], []
        for i in range(1, nrows):
            pos  = scan_sheet.row_values(i)[POS_IDX]
            stat = scan_sheet.row_values(i)[STAT_IDX]
            bc   = scan_sheet.row_values(i)[BC_IDX]

            # 把数字条码转成不带 .0 的字符串
            if isinstance(bc, float):
                bc = str(int(bc))
            else:
                bc = str(bc).strip()

            w    = scan_sheet.row_values(i)[WARM_IDX] if WARM_IDX is not None else ""
            Position.append(pos); Status.append(stat); OriginBarcode.append(bc); Warm.append(w)
            parts = str(bc).split("-", 1)
            CutBarcode.append(parts[0]); SubBarcode.append("-" + parts[1] if len(parts) == 2 else "")

        # 从 Warm 取 Xn
        plate_no_int = 1
        for w in Warm:
            s = str(w)
            if s.upper().startswith("X"):
                m = re.search(r"X(\d+)", s.upper())
                if m:
                    plate_no_int = int(m.group(1))
                    break
        plate_no_str = f"X{plate_no_int}"

        aligned = {
            "Position": Position, "Status": Status, "OriginBarcode": OriginBarcode,
            "CutBarcode": CutBarcode, "SubBarcode": SubBarcode, "Warm": Warm
        }
        plates_payload.append(build_one_plate_payload(aligned, layout='vertical', plate_no_str=plate_no_str))

    else:
        # —— Starlet：支持单板/多板 —— #
        plate_groups = _starlet_split_plates(scan_sheet, scan_index)  # [(n, [rows...]), ...]
        for plate_no_int, row_indexes in plate_groups:
            aligned = _process_one_starlet_plate(scan_sheet, scan_index, row_indexes, plate_no_int)
            plates_payload.append(build_one_plate_payload(aligned, layout='horizontal', plate_no_str=aligned["plate_no_str"]))

    # ========== 4. 保存 Session 与渲染 ==========
    request.session["export_payload"] = {
        "project_name": project_name,
        "project_name_full": project_name_full,
        "instrument_num": instrument_num,
        "systerm_num": systerm_num,
        "platform": platform,
        "injection_plate": injection_plate,
        "today_str": today_str,
        "testing_day": testing_day,
        "plates": plates_payload,                 # ⭐ 多板/单板统一
    }
    request.session.modified = True

    return render(request, "dashboard/ProcessResult.html", {
        "project_name": project_name,
        "project_name_full": project_name_full,
        "platform": platform,
        "today_str": today_str,

        # 模板循环多张卡片
        "plates": plates_payload,
        "station_save_summary": station_save_summary,                 
    })


# preview_payload
def preview_export(request):
    all_payload = request.session.get("export_payload")
    if not all_payload:
        return HttpResponseBadRequest("没有可预览的数据，请先生成结果页面。")

    # 兼容：多板/单板
    if "plates" in all_payload:
        try:
            idx = int(request.GET.get("plate", "0"))
        except ValueError:
            idx = 0
        plates = all_payload["plates"]
        if idx < 0 or idx >= len(plates):
            return HttpResponseBadRequest("板索引无效。")
        payload = plates[idx]   # ⭐ 这里面才有 worksheet_table / error_rows / txt_headers / worklist_records
    else:
        payload = all_payload   # 旧结构，仍包含 worksheet_table 等顶层字段

    nums = [str(i) for i in range(1, 13)]
    project_name = str(all_payload.get("project_name", "PROJECT"))
    project_name_full = str(all_payload.get("project_name_full", "PROJECT"))
    platform = str(all_payload.get("platform", "NewPlatform"))
    return render(
        request,
        "dashboard/export_pdf.html",
        {
            "preview": True,
            "nums": nums,
            "project_name_full": project_name_full,
            "platform": platform,
            **payload,   # 必须把 worksheet_table / error_rows / txt_headers / worklist_records / header 带进去
        },
    )

# 导出pdf时，保存一份json文件用于存储payload中的数据，便于后续替换功能生成新的工作清单
def _json_default(obj):
    """让 json.dump 遇到不可序列化对象时兜底为字符串。"""
    try:
        # datetime/date 等常见类型优先转 ISO
        if hasattr(obj, "isoformat"):
            return obj.isoformat()
    except Exception:
        pass
    return str(obj)


def sanitize_payload(obj):
    """
    递归清理 payload 中的 NaN/Infinity 值
    将其转换为合适的默认值
    """
    if isinstance(obj, dict):
        return {k: sanitize_payload(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_payload(item) for item in obj]
    elif isinstance(obj, float):
        if math.isnan(obj):
            return ""  # 或返回 ""
        if math.isinf(obj):
            return ""  # 或返回 ""
        return obj
    elif isinstance(obj, str):
        if obj.lower() in ("nan", "none"):
            return ""
        return obj
    else:
        return obj
    
def _dump_payload_json(target_dir: str, payload: dict, filename: str = "payload.json") -> str:
    """
    将 payload 保存为 JSON 文件（UTF-8，保留中文），与 PDF 同目录。
    为避免写一半中断，使用 .tmp 原子替换。
    返回写入的绝对路径。
    """
    os.makedirs(target_dir, exist_ok=True)
    out_path = os.path.join(target_dir, filename)
    tmp_path = out_path + ".tmp"

    # ⭐ 清理 NaN 值
    clean_payload = sanitize_payload(payload)

    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=_json_default, allow_nan=False)

    os.replace(tmp_path, out_path)
    return out_path


# 导出pdf和excel
def export_files(request):
    """使用WeasyPrint生成PDF（兼容单板/多板）"""
    all_payload = request.session.get("export_payload")

    if not all_payload:
        return HttpResponseBadRequest("没有可导出的数据，请先生成结果页面。")

    # === 兼容两种会话结构：单板 vs 多板 ===
    # 单板（旧结构）：直接就是一份 payload，包含 worksheet_table 等键
    # 多板（新结构）：export_payload 里有 plates: [ {worksheet_table,..., header{plate_no}}, ... ]
    if "plates" in all_payload and isinstance(all_payload["plates"], list):
        # 读取前端传入的板索引（默认 0）
        try:
            plate_idx = int(request.GET.get("plate", "0"))
        except ValueError:
            plate_idx = 0
        if plate_idx < 0 or plate_idx >= len(all_payload["plates"]):
            return HttpResponseBadRequest("板索引无效。")
        payload = all_payload["plates"][plate_idx]
        # 这些顶层字段依然从总的 all_payload 里取（沿用旧有逻辑）
        payload["project_name"] = all_payload.get("project_name")
        payload["platform"] = all_payload.get("platform")

        if payload["platform"]!="Tecan":
            payload["testing_day"] = all_payload.get("testing_day")
        else:
            pass

    else:
        payload = all_payload  # 旧结构：单板

    # 1) 目录设置
    if payload["platform"]!="Tecan" and payload["platform"]!="手工取样":
        if payload["testing_day"] == "today":
            today_str  = timezone.localtime().strftime("%Y-%m-%d")
        else:
            today_str = (timezone.localtime() + timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        today_str = datetime.today().strftime("%Y-%m-%d")

    project = str(all_payload.get("project_name", ""))
    project_name_full = str(all_payload.get("project_name_full", ""))
    instrument_num = str(all_payload.get("instrument_num", ""))
    systerm_num = str(all_payload.get("systerm_num", ""))
    platform = str(payload.get("platform", "NewPlatform"))
    base_dir = settings.DOWNLOAD_ROOT

    if platform == 'NIMBUS' or platform == 'Tecan' or platform == '手工取样':
        target_dir = os.path.join(base_dir, platform, today_str, project)
    else:
        target_dir = os.path.join(base_dir, platform,'工作清单和上机列表',today_str, project)

    os.makedirs(target_dir, exist_ok=True)

    # 2) 字体路径设置
    font_path = os.path.join(settings.BASE_DIR, 'dashboard', 'static', 'css', 'fonts', 'NotoSansSC-Regular.ttf')

    # 验证字体文件存在
    if not os.path.exists(font_path):
        return JsonResponse({"ok": False, "message": f"字体文件不存在: {font_path}"})

    # 3) 准备模板数据
    nums = [str(i) for i in range(1, 13)]

    # 4) 渲染HTML
    pdf_html = render_to_string(
        "dashboard/export_pdf.html",
        {
            "worksheet_table": payload["worksheet_table"],
            "error_rows": payload["error_rows"],
            "project_name_full": project_name_full,
            "nums": nums,
            "preview": False,  # PDF模式，不使用浏览器字体加载
            "header": payload.get("header", {}), 
            "platform": payload.get("platform", ""),   # ★ 新增：让模板能识别 Starlet
        },
    )

    # 5) 创建字体配置 - WeasyPrint需要
    font_config = FontConfiguration()

    # 6) 定义PDF专用CSS，包含字体配置
    pdf_css = CSS(string=f"""
        @font-face {{
            font-family: "NotoSans";
            src: url("{font_path}") format("truetype");
            font-weight: normal;
            font-style: normal;
        }}
        
        @page {{
            size: A4 landscape;
            margin: 6mm 6mm 8mm 6mm;
        }}
        
        body, table, td, th, div, span {{
            font-family: "NotoSans", "DejaVu Sans", sans-serif;
            font-size: 9pt;
            line-height: 1.25;
        }}
    """, font_config=font_config)

    # 7) 生成PDF
    header = payload.get("header") or {}
    plate_no = header.get("plate_no", "") 
    plate_suffix = f"{plate_no}" if str(plate_no) else ""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    pdf_fname = f"{plate_suffix}_WorkSheet_{instrument_num}_{systerm_num}_{project}_{timestamp}_{plate_suffix}_GZ.pdf"
    pdf_path = os.path.join(target_dir, pdf_fname)

    payload_name = f"{Path(pdf_fname).stem}.payload.json"
    
    try:
        # 使用WeasyPrint生成PDF
        HTML(string=pdf_html).write_pdf(
            pdf_path,
            stylesheets=[pdf_css],
            font_config=font_config
        )

        # ⭐ 新增：首次生成工作清单时，同时落盘保存 payload.json（与PDF同目录）
        _dump_payload_json(target_dir, payload, filename=payload_name)

    except Exception as e:
        return JsonResponse({"ok": False, "message": f"PDF生成失败: {str(e)}"})

    # 8) 生成Excel
    xlsx_path = os.path.join(target_dir, f"上机列表_{timestamp}.xlsx")
    
    # 组装 DataFrame
    headers = payload["txt_headers"]
    records = payload["worklist_records"]  # list[dict]
    df = pd.DataFrame(records, columns=headers)

    # 从会话 payload 里取仪器编号，再查仪器厂家
    instrument_num = (payload.get("header") or {}).get("instrument_num")  # 你在 payload['header'] 里放过它
    instrument_name = ""
    if instrument_num:
        cfg = InstrumentConfiguration.objects.filter(instrument_num=instrument_num).first()
        if cfg:
            instrument_name = (cfg.instrument_name or "").strip()

    worklist_url_key = None  # 返回 JSON 用
    worklist_url_val = None

    if instrument_name.lower() == "sciex":
        # Sciex：导出制表符分隔的 .txt
        # txt_fname = f"OnboardingList_{timestamp}{plate_suffix}.txt"
        txt_fname = f"{plate_suffix}_OnboardingList_{instrument_num}_{systerm_num}_{project}_{timestamp}_{plate_suffix}_GZ.txt"
        txt_path = os.path.join(target_dir, txt_fname)
        df.to_csv(txt_path, sep="\t", index=False, encoding="utf-8")
        worklist_url_key = "txt_url"
        worklist_url_val = f"{settings.DOWNLOAD_URL}{today_str}/{project}/{txt_fname}"

    elif instrument_name.lower() == "thermo" or instrument_name.lower() == "agilent":
        # thermo和agilent：导出逗号分隔的 .csv
        csv_fname = f"{plate_suffix}_OnboardingList_{instrument_num}_{systerm_num}_{project}_{timestamp}_{plate_suffix}_GZ.csv"
        csv_path = os.path.join(target_dir, csv_fname)
        df.to_csv(csv_path, sep=",", index=False, encoding="gbk")
        worklist_url_key = "csv_url"
        worklist_url_val = f"{settings.DOWNLOAD_URL}{today_str}/{project}/{csv_fname}"

    else:
        # 其它厂家：维持原有 .xlsx
        xlsx_fname = f"{plate_suffix}_OnboardingList_{instrument_num}_{systerm_num}_{project}_{timestamp}_{plate_suffix}_GZ.xlsx"
        xlsx_path = os.path.join(target_dir, xlsx_fname)
        with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Worklist", index=False)
        worklist_url_key = "xlsx_url"
        worklist_url_val = f"{settings.DOWNLOAD_URL}{today_str}/{project}/{xlsx_fname}"

    # 返回结果：保留 pdf_url，并根据厂家返回 txt_url 或 xlsx_url
    resp = {
        "ok": True,
        "message": "导出完成",
        "pdf_url": f"{settings.DOWNLOAD_URL}{today_str}/{project}/工作清单_{timestamp}_{plate_suffix}.pdf",
    }
    resp[worklist_url_key] = worklist_url_val
    return JsonResponse(resp)



# 历史标本查找
def sample_search_api(request):
    query = request.GET.get("q", "").strip()
    if not query:
        return JsonResponse({"results": []})

    records = SampleRecord.objects.filter(
        models.Q(sample_name__icontains=query) | models.Q(barcode__icontains=query)
    ).order_by("-record_date")

    results = [
        {
            "project_name": r.project_name,
            "date": r.record_date.strftime("%Y-%m-%d"),
            "plate_no": r.plate_no,
            "well_str": r.well_str,
            "error_info": r.error_info or "",
        }
        for r in records
    ]

    return JsonResponse({"results": results})


# 手工取样
from openpyxl import load_workbook
def Manual_process_result(request):
    """
    手工取样：仅根据板数渲染“曲线+质控”
    - 起始/结束实验号：先收下（本阶段不渲染临床样本）
    - 板数：渲染几块 96 孔板
    - 填充顺序：按 SamplingConfiguration.layout（horizontal/vertical）
    - 别名：若 mapping_file 设置了别名，填入工作清单对应位置
    """
    if request.method != "POST":
        return HttpResponseBadRequest("仅支持POST")

    method_type = request.POST.get("method_type")
    ic(method_type)

    if method_type == "icpms":  # ICP-MS特殊方法逻辑
        ctx = _build_icpms_manual_worksheets(request)

        # ★ 新增：把手工 ICP-MS 生成的结果写入 session，结构与 ProcessResult 保持一致
        header_meta = {
            "today_str":         ctx.get("today_str"),
            "instrument_num":    ctx.get("instrument_num"),
            "injection_plate":   ctx.get("injection_plate"), 
        }

        request.session["export_payload"] = {
            "project_name":      ctx.get("project_name"),
            "project_name_full": ctx.get("project_name_full"),
            "instrument_num":    ctx.get("instrument_num"),
            "systerm_num":       ctx.get("systerm_num"),
            "platform":          ctx.get("platform"),
            "injection_plate":   ctx.get("injection_plate"), 
            "today_str":         ctx.get("today_str"),

            # 手工这里只支持“今天”，直接写 today，就能复用 export_files 里对 testing_day 的逻辑
            # "testing_day":       ctx.get("testing_day", "today"),
            # 多板/单板统一放在 plates 里，preview_export/export_files 都是读这个字段
            "plates":            ctx.get("plates", []),

            "header": header_meta,
        }
        request.session.modified = True
        # ★ END 新增

        return render(request, "dashboard/ProcessResult_Manual.html", ctx)

    # “常规方法”逻辑
    project_id   = request.POST.get("project_id", "").strip()
    start_no     = request.POST.get("start_no", "").strip()
    end_no       = request.POST.get("end_no", "").strip()
    plate_count  = int(request.POST.get("plate_count", "1") or 1)

    # 取项目配置
    cfg = SamplingConfiguration.objects.get(pk=project_id)

    curve_points = int(cfg.curve_points or 0)         # 曲线数
    qc_groups    = int(cfg.qc_groups or 0)            # 质控组数
    qc_levels    = int(cfg.qc_levels or 0)            # 每组水平数
    layout       = (cfg.layout or "horizontal")       # horizontal / vertical
    proj_full    = cfg.project_name_full or cfg.project_name

    # ------- 可选：读取 mapping_file 中的“别名” -------
    # 约定：Excel 第一张表，包含 Name / Alias 两列（如已配置）
    alias_map = {}
    if cfg.mapping_file:
        try:
            wb = load_workbook(cfg.mapping_file.path, data_only=True)
            ws = wb.active
            headers = { (cell.value or "").strip(): cell.column for cell in ws[1] }
            name_col  = headers.get("Name")
            alias_col = headers.get("Alias")
            if name_col and alias_col:
                for r in ws.iter_rows(min_row=2, values_only=True):
                    name  = str(r[name_col-1]).strip() if r[name_col-1] else ""
                    alias = str(r[alias_col-1]).strip() if r[alias_col-1] else ""
                    if name and alias:
                        alias_map[name] = alias
        except Exception:
            alias_map = {}

    # ------- 生成 96 孔网格骨架 -------
    letters = list("ABCDEFGH")
    nums    = list(range(1, 13))  # 1..12
    today_str = timezone.localtime().strftime("%Y-%m-%d")

    # 生成“曲线+质控”的占位清单（线性序列）
    # 约定：前 curve_points 个为“曲线”，之后为 QC（qc_groups * qc_levels）
    slots = []
    # 曲线
    for i in range(1, curve_points + 1):
        name = f"STD-{i}"
        slots.append({
            "type": "STD",
            "name": alias_map.get(name, name),
        })
    # 质控
    for g in range(1, qc_groups + 1):
        for lv in range(1, qc_levels + 1):
            name = f"QC{g}-L{lv}"
            slots.append({
                "type": "QC",
                "name": alias_map.get(name, name),
            })

    # 把线性 slots 按“横向/纵向”塞进 96 孔网格
    def make_empty_plate():
        return [[{
            "letter": letters[r],
            "num": c+1,
            "index": r*12 + c + 1,   # 1..96
            "locator": False,        # 与 Tecan 模板字段兼容
            "locator_warm": "",
            "match_sample": "",      # 兼容模板
            "cut_barcode": "",
            "sub_barcode": "",
            "warm": "",
            "status": "",
            "dup_barcode": "",
            "dup_barcode_sample": "",
        } for c in range(12)] for r in range(8)]

    def fill_plate(plate):
        # 返回已填入“曲线+质控”的 plate（不足 96 个则只填前面）
        seq = list(slots)  # 只填这些
        if layout == "vertical":   # 纵向：A1,A2,...,A12,B1,B2,...
            for c in range(12):
                for r in range(8):
                    if not seq: return
                    cell = plate[r][c]
                    cell["match_sample"] = seq[0]["name"]
                    seq.pop(0)
        else:                      # 横向：A1,B1,...,H1,A2,B2,...
            for r in range(8):
                for c in range(12):
                    if not seq: return
                    cell = plate[r][c]
                    cell["match_sample"] = seq[0]["name"]
                    seq.pop(0)

    worksheet_list = []  # 多板
    for p in range(plate_count):
        plate = make_empty_plate()
        fill_plate(plate)
        worksheet_list.append(plate)

    # —— 为了最大限度复用 Tecan 的结果模板字段命名 —— 
    # （plate_number/nums/worksheet_table 等，见 Tecan 结果页模板变量） :contentReference[oaicite:3]{index=3}
    ctx = {
        "platform": "Manual",
        "project_name_full": proj_full,
        "today_str": today_str,
        "nums": nums,
        "plate_number": plate_count,
        "plates": worksheet_list,         # 多板
    }

    # 采用独立模板（仅渲染工作清单，无上机列表）
    return render(request, "dashboard/ProcessResult_Manual.html", ctx)


def _build_icpms_manual_worksheets(request):
    """
    手工取样模块 - ICP-MS 特殊方法
    生成：
      - 多板工作清单 worksheet_table
      - 报错信息表 error_rows（仅 No match）
      - 上机列表 worklist_records / txt_headers

    前端需上传：
      - station_list: 岗位清单表（格式与 NIMBUS 相同）
      - scan_result : 扫码结果表（格式为 ICP-MS 手工取样模板，B3:P34 为条码区域，I1 为起始板号）
    """

    project_id     = request.POST.get("project_id", "").strip()
    instrument_num = request.POST.get("instrument_num", "").strip()
    systerm_num    = request.POST.get("systerm_num", "").strip()
    injection_plate = request.POST.get("injection_plate")

    station_file = request.FILES.get("station_list")
    scan_file    = request.FILES.get("scan_result")

    if not (project_id and station_file and scan_file):
        raise ValueError("缺少必要参数或文件（project_id / station_list / scan_result）")

    # 1) 项目配置 & 对应关系表（工作清单 sheet）
    cfg = SamplingConfiguration.objects.get(pk=project_id)
    proj_name = cfg.project_name
    proj_name_full = cfg.project_name_full
    mapping_path   = cfg.mapping_file.path

    df_mapping_wc = pd.read_excel(mapping_path, sheet_name="工作清单")
    # Barcode -> Name / Code，用于识别曲线/QC/Blank
    barcode_to_name = dict(
        zip(df_mapping_wc["Barcode"].astype(str), df_mapping_wc["Name"].astype(str))
    )
    barcode_to_code = dict(
        zip(df_mapping_wc["Barcode"].astype(str), df_mapping_wc["Code"].astype(str))
    )

    # 2) 岗位清单：主条码 -> 实验号 列表（与 NIMBUS 完全同源）
    station_book = xlrd.open_workbook(filename=None, file_contents=station_file.read())
    st_sheet   = station_book.sheets()[0]
    st_nrows   = st_sheet.nrows
    st_ncols   = st_sheet.ncols
    st_header  = [str(st_sheet.row_values(0)[i]).strip() for i in range(st_ncols)]
    st_index   = {col: idx for idx, col in enumerate(st_header)}

    MB_IDX = st_index.get("主条码", 0)
    SN_IDX = st_index.get("实验号", 0)

    barcode_to_names = defaultdict(list)
    for i in range(1, st_nrows):
        bc = str(st_sheet.row_values(i)[MB_IDX]).strip()
        sn = str(st_sheet.row_values(i)[SN_IDX]).strip()
        if bc:
            barcode_to_names[bc].append(sn)

    # 3) 扫码结果表：读取 B3:P34 中的条码，按列纵向收集为列表 A
    wb_scan = load_workbook(filename=scan_file, data_only=True)
    ws_scan = wb_scan.active

    list_a = []  # 条码列表 A
    for col in range(2, 17):      # B(2) .. P(16)
        for row in range(3, 35):  # 3 .. 34
            val = ws_scan.cell(row=row, column=col).value
            if val is None:
                continue
            s = str(val).strip()
            if s:
                list_a.append(s)

    # 起始板号：I1 单元格中的数字
    start_plate_no = 1
    raw_plate = ws_scan["I1"].value
    if raw_plate is not None:
        m = re.search(r"(\d+)", str(raw_plate))
        if m:
            start_plate_no = int(m.group(1))

    # 4) 根据列表 A 生成 cut_barcode 列表，并按 NIMBUS 规则匹配实验号
    cut_barcodes = []
    origin_barcodes = []
    for bc in list_a:
        bc_str = str(bc).strip()
        parts = bc_str.split("-", 1)
        cut = parts[0]
        cut_barcodes.append(cut)
        origin_barcodes.append(bc_str)

    cut_counter = Counter(cut_barcodes)

    match_sample_raw   = []  # 与 NIMBUS 中 MatchSampleName 同源（未经过 mapping 表转名称）
    match_result       = []  # TRUE/FALSE
    dup_barcode        = []
    dup_barcode_sample = []

    for cb in cut_barcodes:
        cb_str = str(cb)
        matched_names = barcode_to_names.get(cb_str, [])
        cb_count = cut_counter[cb_str]

        if len(matched_names) == 1:
            match_result.append("TRUE")
            match_sample_raw.append(matched_names[0])
            dup_barcode.append("")
            dup_barcode_sample.append("")

        elif len(matched_names) == 0:
            match_result.append("FALSE")
            if cb_str != "":
                match_sample_raw.append(cb_str)
            else:
                match_sample_raw.append("")
            dup_barcode.append("")
            dup_barcode_sample.append("")

        elif len(matched_names) == 2:
            match_result.append("TRUE")
            if matched_names[0] == matched_names[1]:
                dup_barcode.append("Likely")
                match_sample_raw.append(matched_names[0])
                dup_barcode_sample.append("")
            else:
                dup_barcode.append("TRUE")
                match_sample_raw.append(matched_names[0] + "-" + matched_names[1])
                dup_barcode_sample.append("TRUE" if cb_count >= 2 else "")

        else:
            match_result.append("TRUE")
            unique_Middlelist = list(dict.fromkeys(matched_names))
            order = {'VF': 1, 'AE': 2, 'VD': 3, 'V': 4, 'VK': 5, 'WV': 6}

            def sort_key(x):
                for prefix_length in (2, 1):
                    prefix = x[:prefix_length]
                    if prefix in order:
                        return order[prefix]
                return len(order) + 1

            sorted_lis = sorted(unique_Middlelist, key=sort_key)
            if len(unique_Middlelist) >= 2:
                match_sample_raw.append('-'.join(sorted_lis))
                dup_barcode.append("TRUE")
                dup_barcode_sample.append("TRUE" if cb_count >= 2 else "")
            else:
                match_sample_raw.append(matched_names[0])
                dup_barcode.append("")
                dup_barcode_sample.append("")

    # 结合 df_mapping_wc，把 TRUE/ FALSE 的 raw 值转换成最终实验号 / 曲线名
    list_b_items = []  # 每个元素：{"barcode":cut,"origin_barcode":..., "sample_name":..., "is_special":bool}
    for i, cb in enumerate(cut_barcodes):
        raw_value = str(match_sample_raw[i])
        if match_result[i] == "TRUE":
            sample_name = raw_value
        else:
            if raw_value == "":
                sample_name = ""
            else:
                # 在 mapping_file 中找曲线 / 质控名称
                sample_name = barcode_to_name.get(raw_value, "No match")

        code = barcode_to_code.get(str(cb), "")
        is_special = bool(code) and (
            str(code).startswith("STD")
            or str(code).startswith("QC")
            or str(code).startswith("Blank")
        )

        list_b_items.append({
            "barcode": str(cb),
            "origin_barcode": origin_barcodes[i],
            "sample_name": sample_name,
            "is_special": is_special,
            "dup_barcode": dup_barcode[i],
            "dup_barcode_sample": dup_barcode_sample[i],
        })

    # 5) 提取“曲线+质控”模板（每块板都要重复）和临床样本队列
    special_pattern = []
    seen_special_bc = set()
    clinical_queue  = []

    for item in list_b_items:
        if item["is_special"]:
            if item["barcode"] not in seen_special_bc:
                seen_special_bc.add(item["barcode"])
                special_pattern.append(item)
        else:
            clinical_queue.append(item)

    # 每块板的可用孔位数：96 孔 - 1 个 H12 比对孔 - 1 个定位孔
    per_plate_capacity = 96 - 2

    # 每块板留给临床样本的孔数（总容量减去曲线/质控占用）
    clin_per_plate = max(per_plate_capacity - len(special_pattern), 0)
    plate_count = max(1, ceil(len(clinical_queue) / clin_per_plate))

    # 6) 生成多板工作清单（纵向填充；A3/B3/C3... 为定位孔；每块板 H12 为空）
    letters = list("ABCDEFGH")
    nums    = [str(i) for i in range(1, 13)]
    today_str = timezone.localtime().strftime("%Y-%m-%d")

    def make_empty_plate():
        grid = []
        for r, row_letter in enumerate(letters):
            row = []
            for c, col_num in enumerate(nums):
                well_str = f"{row_letter}{col_num}"
                index    = r * 12 + int(col_num)
                row.append({
                    "letter": row_letter,
                    "num": col_num,
                    "well_str": well_str,
                    "index": index,
                    "locator": False,
                    "locator_warm": "",
                    "match_sample": "",
                    "cut_barcode": "",
                    "sub_barcode": "",
                    "origin_barcode": "",
                    "warm": "",
                    "status": "",
                    "dup_barcode": "",
                    "dup_barcode_sample": "",
                })
            grid.append(row)
        return grid

    def iter_fill_positions(skip_row, skip_col):
        """
        纵向填充顺序生成 (row_idx, col_idx)，
        跳过：H12（比对孔）以及 (skip_row, skip_col) 定位孔。
        """
        for c in range(12):         # 0..11 -> 列 1..12
            for r in range(8):      # 0..7  -> 行 A..H
                # 跳过 H12
                if r == 7 and c == 11:
                    continue
                # 跳过定位孔
                if r == skip_row and c == skip_col:
                    continue
                yield r, c

    plates = []
    clinical_idx = 0

    for p in range(plate_count):
        plate = make_empty_plate()
        error_rows_plate = []   # 本块板的报错信息

        # 定位孔：板1=A3, 板2=B3, 板3=C3...
        plate_no_int = start_plate_no + p
        plate_no_str = f"{plate_no_int}"

        # 板1=A3, 板2=B3, 板3=C3... → 行号 = (板号 - 1)
        locator_row = (plate_no_int - 1) % 8       # 0=A,1=B,...,7=H，超过8块再循环
        locator_col = 2                            # 第3列 → “3”这一列

        # 先计算填样顺序（会跳过定位孔和 H12）
        pos_iter = iter(iter_fill_positions(locator_row, locator_col))

        if locator_row < 8:
            loc_cell = plate[locator_row][locator_col]
            loc_cell["locator"] = True
            loc_cell["locator_warm"] = f"X{plate_no_str}"

        # 先填“曲线+质控”固定模式（每块板复用同一批条码/实验号）
        for item in special_pattern:
            try:
                r, c = next(pos_iter)
            except StopIteration:
                break
            cell = plate[r][c]
            cell["match_sample"]      = item["sample_name"]
            cell["origin_barcode"]    = item["origin_barcode"]
            cell["cut_barcode"]       = item["barcode"]
            parts = str(item["origin_barcode"]).split("-", 1)
            cell["sub_barcode"]       = "-" + parts[1] if len(parts) == 2 else ""
            cell["dup_barcode"]       = item["dup_barcode"]
            cell["dup_barcode_sample"]= item["dup_barcode_sample"]

            # 报错信息：只关心 match_sample == "No match"
            if cell["match_sample"] == "No match":
                error_rows_plate.append({
                    "sample_name": cell["match_sample"],
                    "origin_barcode": cell["origin_barcode"],
                    "plate_no": plate_no_str,
                    "well_str": cell["well_str"],
                    "warn_level": "",
                    "warn_info": "",
                })

        # 再填临床样本
        filled_clin_this_plate = 0
        while clinical_idx < len(clinical_queue) and filled_clin_this_plate < clin_per_plate:
            try:
                r, c = next(pos_iter)
            except StopIteration:
                break

            item = clinical_queue[clinical_idx]
            clinical_idx += 1
            filled_clin_this_plate += 1

            cell = plate[r][c]
            cell["match_sample"]      = item["sample_name"]
            cell["origin_barcode"]    = item["origin_barcode"]
            cell["cut_barcode"]       = item["barcode"]
            parts = str(item["origin_barcode"]).split("-", 1)
            cell["sub_barcode"]       = "-" + parts[1] if len(parts) == 2 else ""
            cell["dup_barcode"]       = item["dup_barcode"]
            cell["dup_barcode_sample"]= item["dup_barcode_sample"]

            # 报错信息：只关心 match_sample == "No match"
            if cell["match_sample"] == "No match":
                error_rows_plate.append({
                    "sample_name": cell["match_sample"],
                    "origin_barcode": cell["origin_barcode"],
                    "plate_no": plate_no_str,
                    "well_str": cell["well_str"],
                    "warn_level": "",
                    "warn_info": "",
                })

        plates.append({
            "plate_no": plate_no_str,
            "worksheet_table": plate,
            "error_rows": error_rows_plate,
        })

        if clinical_idx >= len(clinical_queue):
            break

    ############################################
    # 7) 上机列表生成（与 NIMBUS 逻辑基本一致）
    ############################################

    # 对应关系表：上机列表 sheet
    df_worklistmap = pd.read_excel(mapping_path, sheet_name="上机列表").fillna("")

    # 仪器上机模板 → 确定列头
    instrument_config = InstrumentConfiguration.objects.get(
        instrument_num=instrument_num,
        systerm_num=systerm_num
    )
    raw = instrument_config.upload_file.read()
    try:
        text = raw.decode("utf-8")
    except Exception:
        text = raw.decode("gbk", errors="replace")

    df_template = pd.read_csv(
        StringIO(text),
        sep=None,
        engine="python",
        dtype=str
    )
    txt_headers = df_template.columns.tolist()

    # 仪器名称（Thermo / Agilent 等）
    instrument_name = getattr(instrument_config, "instrument_name", "")

    # 曲线 / QC 名称集合
    test_count   = cfg.test_count or 0
    curve_points = cfg.curve_points or 0

    df_std = df_mapping_wc[
        df_mapping_wc["Code"].astype(str).str.startswith("STD")
    ].copy()
    std_names = df_std["Name"].tolist()
    std_names_use = std_names[: curve_points + 1]

    qc_names = df_mapping_wc[
        df_mapping_wc["Code"].astype(str).str.startswith("QC")
    ]["Name"].astype(str).unique().tolist()

    # 对每块板分别生成上机列表
    for p in plates:
        plate_no_str = p["plate_no"]

        # ★ 每块板单独构建一次 name_to_barcodes（从 df_mapping_wc 拷贝）
        name_to_barcodes = defaultdict(deque)
        for _, row_m in df_mapping_wc.iterrows():
            n = str(row_m.get("Name", "")).strip()
            b = str(row_m.get("Barcode", "")).strip()
            if n and b and b != "nan":
                name_to_barcodes[n].append(b)
        
        ic(name_to_barcodes)

        # ---------- 1) 构建 barcode -> [(well_pos, well_no)...] 队列 ----------
        barcode_to_well = defaultdict(deque)

        rows_grid = p["worksheet_table"]  # 8 行 x 12 列
        for r_idx, row in enumerate(rows_grid):
            for cell in row:
                origin_bc = str(cell["origin_barcode"]).strip()
                if origin_bc and origin_bc != "nan":
                    well_pos = cell["well_str"]          # A1, B5 这种
                    well_no  = (r_idx * 12) + int(cell["num"])  # 1..96 行优先编号
                    barcode_to_well[origin_bc].append((well_pos, well_no))

        # ---------- 2) 构造第一列 SampleName：DB/Test + STD/QC + 临床条码 + QC ----------
        test_list   = ["DB1"] + [f"Test{i}" for i in range(1, test_count + 1)]
        curve_list  = std_names_use
        qc_list1    = ["DB1"] + qc_names
        qc_list2    = qc_names

        # ★ 临床样本：从工作清单中“纵向”遍历，取 **条码** 而不是实验号
        clinical_list = []
        for c in range(12):         # 列优先：A1,A2,...,H12,B1,B2,...
            for r in range(8):
                cell = rows_grid[r][c]
                name = str(cell["match_sample"]).strip()
                if not name:
                    continue
                if name == "No match":
                    continue
                if name in std_names_use or name in qc_names:
                    continue
                if name.startswith("DB") or name.startswith("Test") or name.startswith("Blank"):
                    continue
                origin_bc = str(cell["origin_barcode"]).strip()
                if not origin_bc:
                    continue
                clinical_list.append(origin_bc)      # ★ 第一列使用条码

        SampleName_list = curve_list + qc_list1 + clinical_list + qc_list2
        SampleName_list = [
            x for x in SampleName_list
            if isinstance(x, str) and x
        ]

        # 3) 建立空 worklist DataFrame
        df_worklist = pd.DataFrame(columns=txt_headers)
        first_col_header = txt_headers[0]
        df_worklist[first_col_header] = SampleName_list

        mirror_cols = set()

        # 4) 应用 df_worklistmap 映射规则
        col0 = df_worklist[first_col_header]

        def resolve_vialpos_for_value(sample_name_value, placeholder):
            """
            根据第一列的值 sample_name_value 决定孔位。
            - 当 placeholder 为 {{Well_Number}} / {{Well_Position}} 时，按条码动态计算；
            - 否则视为固定字符串（例如 D3B-F8 / D3B-F9），直接返回 placeholder。
            """

            s = str(sample_name_value).strip()
            if not s:
                return ""

            # 1) 非占位符：直接返回配置值（用于 DB* / Test* 的 D3B-F8 / D3B-F9）
            if placeholder not in ("{{Well_Number}}", "{{Well_Position}}"):
                return placeholder

            # 1) QC/STD：通过 name_to_barcodes 队列取条码，再查位置信息
            if s in name_to_barcodes and name_to_barcodes[s]:
                barcode = name_to_barcodes[s].popleft()
                wells_q = barcode_to_well.get(barcode)
                if wells_q:
                    pos, no = wells_q.popleft()
                    # ICP-MS 这里不区分 Thermo/Agilent，格式保持与 NIMBUS 一致
                    if placeholder == "{{Well_Number}}":
                        return f"{injection_plate}:{no}"
                    else:
                        return f"{injection_plate}-{pos}"

            # 2) 临床样本：第一列即为条码
            if s in barcode_to_well:
                wells_q = barcode_to_well.get(s)
                if wells_q:
                    pos, no = wells_q.popleft()
                    if placeholder == "{{Well_Number}}":
                        return f"{injection_plate}:{no}"
                    else:
                        return f"{injection_plate}-{pos}"

            # 3) 找不到就返回空
            return f"{injection_plate}-{'H1'}"

        def apply_to(mask, fill_values):
            nonlocal mirror_cols
            for col, val in zip(txt_headers[1:], fill_values.values):
                v = str(val).strip()

                # "*"：该列镜像第一列
                if v == "*":
                    mirror_cols.add(col)
                    continue

                # 2) 进样体积列：直接用配置的 injection_volume（如果有），否则写回映射表里的值
                if col in ("SmplInjVol", "Injection volume"):
                    df_worklist.loc[mask, col] = injection_volume or v
                    continue

                # 3) ★ 动态计算孔位列：列名为 VialPos / Vial position / 样品瓶
                #    触发方式与 NIMBUS 完全一致，不再通过占位符 "{{Well_Number}}" 判断
                if col in ("VialPos", "Vial position", "样品瓶", "样品瓶号"):
                    df_worklist.loc[mask, col] = df_worklist.loc[mask, first_col_header].apply(
                        lambda x: resolve_vialpos_for_value(x, v)
                    )
                    continue

                # 4) 其它列：直接按映射表填固定值
                df_worklist.loc[mask, col] = val

        # 遍历 mapping 表中的每一条规则（完全照 NIMBUS 的 key 语义）
        for _, rule in df_worklistmap.iterrows():
            sample_key  = rule.iloc[0]
            fill_values = rule.iloc[1:]

            if str(sample_key) == "DB*":
                mask = df_worklist.iloc[:, 0].str.startswith("DB")
                apply_to(mask, fill_values)
            elif str(sample_key).startswith("DB"):
                mask = col0 == str(sample_key)
                apply_to(mask, fill_values)
            elif str(sample_key) == "Test*":
                mask = col0.str.startswith("Test")
                apply_to(mask, fill_values)
            elif str(sample_key).startswith("Test"):
                mask = col0 == str(sample_key)
                apply_to(mask, fill_values)

            elif str(sample_key) == "STD*":
                mask = col0.isin(std_names_use)
                apply_to(mask, fill_values)
            elif str(sample_key).startswith("STD"):
                mask = col0 == str(sample_key)
                apply_to(mask, fill_values)

            elif str(sample_key) == "QC*":
                mask = col0.isin(qc_names)
                apply_to(mask, fill_values)
            elif str(sample_key).startswith("QC"):
                mask = col0 == str(sample_key)
                apply_to(mask, fill_values)

            elif str(sample_key) == "*":
                # 通配：填充剩余所有空行
                mask = df_worklist.iloc[:, 1].isna() 
                apply_to(mask, fill_values)

        # 5) 镜像列填充
        for col in mirror_cols:
            df_worklist[col] = df_worklist[first_col_header]

        txt_headers = ["跳过", "样品类型", "样品名称", "样品瓶号", "级别", "总稀释倍数"]
    
        p["txt_headers"]      = txt_headers
        p["worklist_records"] = df_worklist.to_dict(orient="records")
    

    # 8) 返回给视图的 payload
    ctx = {
        "instrument_num":instrument_num,
        "systerm_num":systerm_num,
        "injection_plate":injection_plate,
        "platform": "手工取样",
        "project_name": proj_name,
        "project_name_full": proj_name_full,
        "today_str": today_str,
        "nums": nums,
        "plate_number": len(plates),
        "plates": plates,
    }
    return ctx

