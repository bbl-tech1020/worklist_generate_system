from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect
from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from collections import defaultdict, Counter
from datetime import date, timedelta
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
from icecream import ic
import pandas as pd
import xlrd
import xlwt
import os
import re
import json

from .models import *


# ========== ★ 新增：报错关键词列表 ==========
ERROR_KEYWORDS = [
    "无料",
    "有料",
    "取料NG",
    "扫码NG",
    "开盖NG",
    "吸液NG",
    "分液NG",
    "吸液曲线NG",
    "分液曲线NG",
    "吸/分液曲线NG",
    "待回收",
]

def _parse_preprocess_worksheet(file_obj):
    """
    解析仪器输出的"前处理样品工作单"Excel文件。
    固定结构：
      - 数据区域：Excel 行5~12（对应96孔板A~H行），列C~N（对应孔位1~12）
      - 每个单元格内容为多行文本（\n分隔）：
          lines[0]: 孔位序号 + 符号（如 "2 ○"）
          lines[1]: 实验号（如 "H5275"），空孔无此行
          lines[2]: 条码（如 "7602579116-01"），空孔无此行，某些孔可能跨行
    返回：8行×12列的二维列表，每格为 dict。
    """
    import openpyxl
    import openpyxl.reader.excel as oxl_reader
    import re
    import io
    import zipfile

    # ★ 核心修复：读取原始文件字节流，预处理 ZIP 内容
    raw_bytes = file_obj.read()

    # ★ 核心修复：monkey-patch read_custom，让它对解析错误静默跳过
    # 保存原始方法
    _original_read_custom = oxl_reader.ExcelReader.read_custom

    def _safe_read_custom(self):
        try:
            _original_read_custom(self)
        except Exception:
            pass  # 忽略 custom.xml 的任何解析错误

    # 临时替换
    oxl_reader.ExcelReader.read_custom = _safe_read_custom

    try:
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(raw_bytes),
            data_only=True,
            read_only=True,
        )
    finally:
        # ★ 无论成功失败，都还原原始方法，避免影响其他地方的 openpyxl 调用
        oxl_reader.ExcelReader.read_custom = _original_read_custom
        
    ws = wb.active

    letters = list("ABCDEFGH")
    # Excel 行索引：5~12 对应 A~H（openpyxl 行从1开始）
    EXCEL_DATA_ROWS = range(5, 13)   # 5,6,7,8,9,10,11,12
    # Excel 列索引：3~14 对应孔位1~12（C=3, D=4, ..., N=14）
    EXCEL_DATA_COLS = range(3, 15)   # 3..14

    # 预读所有单元格值（read_only 模式下按行读取效率更高）
    # 转成普通二维列表，rows[r][c] = 单元格值（字符串或None）
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=12, min_col=1, max_col=14, values_only=True):
        all_rows.append(list(row))

    table = []
    for r_idx, excel_row in enumerate(EXCEL_DATA_ROWS):
        row_letter = letters[r_idx]   # A, B, ..., H
        row_cells = []
        for c_idx, excel_col in enumerate(EXCEL_DATA_COLS):
            col_num    = c_idx + 1      # 1~12
            well_index = r_idx * 12 + col_num   # 1~96

            raw_value = all_rows[excel_row - 1][excel_col - 1]
            raw_str   = str(raw_value).strip() if raw_value is not None else ""

            # split 换行符，过滤空行
            lines = [l.strip() for l in raw_str.replace("\r\n", "\n").replace("\r", "\n").split("\n") if l.strip()]

            # 解析第一行：序号 + 符号
            symbol = ""
            if lines:
                m = re.match(r"^(\d+)\s*(.*)?$", lines[0])
                if m:
                    symbol = m.group(2).strip()  # ○ / ✗ / ★ 或空

            # 实验号：lines[1]（若存在）
            match_sample = lines[1].strip() if len(lines) >= 2 else ""

            # 条码：lines[2]（若存在）
            # 注意：若条码含换行（如共血条码多行），从lines[2]起拼接
            barcode = ""
            if len(lines) >= 3:
                barcode = "\n".join(lines[2:])

            # 判断是否为空孔（无实验号）
            is_empty = (match_sample == "")

            cell = {
                "letter":       row_letter,
                "num":          str(col_num),
                "well_str":     f"{row_letter}{col_num}",
                "index":        well_index,
                "match_sample": match_sample,
                "barcode":      barcode,
                "symbol":       symbol,   # ○ / ✗ / ★ 等原始符号
                "is_empty":     is_empty,
            }
            row_cells.append(cell)
        table.append(row_cells)

    wb.close()
    return table



# ========== 第一步：处理上传数据，返回中间展示页面 ==========
def WholeBloodWorkstationResult(request):
    """
    全血工作站取样结果处理函数（第一步）
    从取样总表中读取 ScannerCode、Row、Column，生成96孔板工作清单预览
    """
    if request.method != 'POST':
        return HttpResponseBadRequest("仅支持 POST 请求")
    
    # ========== 1. 获取表单参数 ==========
    project_id = request.POST.get("project_id")
    project_name = request.POST.get("project_name")
    platform = "WholeBloodWorkstation"
    instrument_num = request.POST.get("instrument_num")
    systerm_num = request.POST.get("systerm_num")
    testing_day = request.POST.get("testing_day", "today")
    
    # 获取上传的两个文件
    station_list = request.FILES.get('station_list')      # 岗位清单表
    sampling_summary = request.FILES.get('sampling_summary')  # 取样总表
    # ★ 新增：工作清单表
    preprocess_worksheet = request.FILES.get('preprocess_worksheet')
    
    if not (station_list and sampling_summary and project_id and instrument_num):
        return HttpResponseBadRequest("缺少必要参数或文件")

    # ★ 新增：解析工作清单表（可选），生成 worksheet_table_2
    worksheet_table_2 = []
    if preprocess_worksheet:
        worksheet_table_2 = _parse_preprocess_worksheet(preprocess_worksheet)
    
    # 获取项目配置信息
    try:
        config = SamplingConfiguration.objects.get(pk=project_id)
        project_name_full = config.project_name_full
    except:
        project_name_full = project_name
    
    # 确定日期
    if testing_day == "today":
        today_str = timezone.localtime().strftime("%Y%m%d")
        record_date = date.today()
    else:
        today_str = (timezone.localtime() + timedelta(days=1)).strftime("%Y%m%d")
        record_date = date.today() + timedelta(days=1)
    
    # ========== 2. 读取岗位清单表（获取条码→实验号映射）==========
    station_wb = xlrd.open_workbook(filename=None, file_contents=station_list.read())
    station_sheet = station_wb.sheets()[0]
    station_nrows = station_sheet.nrows
    station_ncols = station_sheet.ncols
    
    # 解析表头
    station_header = [str(station_sheet.row_values(0)[i]).strip() for i in range(station_ncols)]
    station_index = {col: idx for idx, col in enumerate(station_header)}
    
    # ★ 与NIMBUS对齐：优先检测是否含"子条码"列，动态决定匹配键类型
    use_sub_barcode = "子条码" in station_index
    if use_sub_barcode:
        KEY_IDX = station_index["子条码"]   # 用完整子条码（如 2437871821-01）作为键
    else:
        KEY_IDX = station_index.get("主条码", 0)  # 退而用主条码（如 2437871821）作为键

    sn_idx = station_index.get("实验号", 0)
    
    # ========== 构建条码→实验号列表的映射 ==========
    barcode_to_names = defaultdict(list)
    for i in range(1, station_nrows):
        barcode = str(station_sheet.row_values(i)[KEY_IDX]).strip()
        sample_name = str(station_sheet.row_values(i)[sn_idx]).strip()
        if barcode and sample_name:
            barcode_to_names[barcode].append(sample_name)
    
    
    # ========== 3. 读取取样总表（从"产品信息"工作表）==========
    summary_wb = xlrd.open_workbook(filename=None, file_contents=sampling_summary.read())
    
    # 查找"产品信息"工作表
    product_sheet = None
    for sheet in summary_wb.sheets():
        if "产品信息" in sheet.name:
            product_sheet = sheet
            break
    
    if product_sheet is None:
        return render(request, "dashboard/error.html", {
            "message": "取样总表中未找到'产品信息'工作表,请检查文件格式"
        })
    
    # 解析产品信息表头
    product_nrows = product_sheet.nrows
    product_ncols = product_sheet.ncols
    product_header = [str(product_sheet.row_values(0)[i]).strip() for i in range(product_ncols)]
    product_index = {col: idx for idx, col in enumerate(product_header)}
    
    # 获取关键列索引
    scanner_code_idx = product_index.get("ScannerCode")
    row_idx = product_index.get("Row")
    column_idx = product_index.get("Column")

    process_no_str_idx = product_index.get("ProcessNoStr")
    
    if scanner_code_idx is None or row_idx is None or column_idx is None:
        return render(request, "dashboard/error.html", {
            "message": f"取样总表'产品信息'工作表中缺少必要列(ScannerCode/Row/Column)\n当前表头: {' | '.join(product_header)}"
        })
    
    # ========== 4. 构建96孔板数据结构 ==========
    letters = list("ABCDEFGH")
    nums = [str(i) for i in range(1, 13)]

    # 用字典暂存：(row, col) -> barcode
    well_data = {}

    # 报错信息
    well_errors = {}
    
    # 读取产品信息工作表数据
    for i in range(1, product_nrows):
        row_val = product_sheet.row_values(i)[row_idx]
        col_val = product_sheet.row_values(i)[column_idx]
        scanner_code = product_sheet.row_values(i)[scanner_code_idx]

        process_no_str = ""
        if process_no_str_idx is not None:
            process_no_str = str(product_sheet.row_values(i)[process_no_str_idx]).strip()
        
        try:
            row_num = int(float(row_val))
            col_num = int(float(col_val))
        except:
            continue
        
        # 验证范围
        if not (1 <= row_num <= 8 and 1 <= col_num <= 12):
            continue
        
        well_data[(row_num, col_num)] = str(scanner_code).strip()

        # ========== ★ 新增：检测 ProcessNoStr 是否包含报错关键词 ==========
        if process_no_str:
            for keyword in ERROR_KEYWORDS:
                if keyword in process_no_str:
                    well_errors[(row_num, col_num)] = process_no_str
                    break  # 匹配到一个关键词即可，无需继续检测
    
    # ========== 5. 按96孔板顺序填充数据 ==========
    worksheet_grid = [[None for _ in nums] for _ in letters]
    error_rows = []
    
    for row_num in range(1, 9):  # 1-8
        for col_num in range(1, 13):  # 1-12
            row_letter = letters[row_num - 1]
            col_str = nums[col_num - 1]
            well_pos = f"{row_letter}{col_str}"
            well_index = (row_num - 1) * 12 + col_num  # 1-96
            
            # 获取该孔位的条码
            barcode = well_data.get((row_num, col_num), "NOTUBE")
            
            # ★ 与NIMBUS对齐：根据 use_sub_barcode 决定切割方式和匹配键
            match_sample = "No match"
            if barcode and barcode != "NOTUBE":
                if use_sub_barcode:
                    # 子条码模式：用原始完整条码（如 2437871821-01）直接匹配
                    cut_barcode = barcode
                    sub_barcode = ""
                    origin_barcode = barcode
                    match_key = barcode
                else:
                    # 主条码模式：切割后取主条码部分匹配
                    parts = barcode.split("-", 1)
                    cut_barcode = parts[0]
                    sub_barcode = "-" + parts[1] if len(parts) == 2 else ""
                    origin_barcode = barcode
                    match_key = cut_barcode
            else:
                cut_barcode = "NOTUBE" if barcode == "NOTUBE" else ""
                sub_barcode = ""
                origin_barcode = barcode if barcode else "NOTUBE"
                match_key = None

            # ★ 与NIMBUS对齐：实验号匹配 + 多实验号拼接展示
            warm = ""
            if match_key and match_key != "NOTUBE":
                matched_names = barcode_to_names.get(match_key, [])
                unique_names = list(dict.fromkeys(matched_names))  # 去重保序
                if len(unique_names) == 1:
                    match_sample = unique_names[0]
                elif len(unique_names) > 1:
                    # ★ 与NIMBUS对齐：多实验号拼接展示 + 共血警告
                    match_sample = "/".join(str(n) for n in unique_names)
                    warm = "共血"
                else:
                    match_sample = "No match"

            # 构建单元格数据
            cell = {
                "letter": row_letter,
                "num": col_str,
                "well_str": well_pos,
                "index": well_index,
                "locator": False,
                "locator_warm": "",
                "match_sample": match_sample,
                "cut_barcode": cut_barcode,
                "sub_barcode": sub_barcode,
                "origin_barcode": barcode,
                "warm": warm,
                "status": "Used",
                "dup_barcode": "",
                "dup_barcode_sample": "",
                "flag_suck": "",
                "flag_dispense": "",
            }
            
            worksheet_grid[row_num - 1][col_num - 1] = cell

            # 收集报错信息
            error_message = well_errors.get((row_num, col_num))  # 从取样总表获取的报错信息
            if error_message:
                # ProcessNoStr 包含报错关键词
                error_rows.append({
                    "sample_name": match_sample if match_sample else barcode,
                    "origin_barcode": barcode,
                    "plate_no": "X1",
                    "well_str": well_pos,
                    "warn_info": error_message,  # 使用 ProcessNoStr 的内容
                })
            elif match_sample == "No match" and barcode != "NOTUBE":
                # 实验号匹配失败
                error_rows.append({
                    "sample_name": match_sample,
                    "origin_barcode": barcode,
                    "plate_no": "X1",
                    "well_str": well_pos,
                    "warn_info": "No match",
                })

    
    # ========== 生成上机列表（针对'全血七项'项目）==========
    worklist_records = []
    txt_headers = []
    
    # 检查是否是'全血七项'项目
    if '全血七项' in project_name or '全血七项' in project_name_full:
        # 查找取样总表中包含'GZKM'关键词的子表
        gzkm_sheet = None
        for sheet in summary_wb.sheets():
            if 'GZKM' in sheet.name:
                gzkm_sheet = sheet
                break
        
        if gzkm_sheet is not None:
            # 读取表格A（GZKM子表）
            gzkm_nrows = gzkm_sheet.nrows
            gzkm_ncols = gzkm_sheet.ncols
            
            # 解析表头
            gzkm_header = [str(gzkm_sheet.row_values(0)[i]).strip() for i in range(gzkm_ncols)]
            txt_headers = gzkm_header
            
            # 获取'样品名称'列索引
            sample_name_col_idx = None
            for idx, col_name in enumerate(gzkm_header):
                if '样品名称' in col_name:
                    sample_name_col_idx = idx
                    break
            
            if sample_name_col_idx is not None:
                # 读取所有数据行，并删除'样品名称'列中含有'SB'关键词的行
                table_b_rows = []
                for i in range(1, gzkm_nrows):
                    row_data = gzkm_sheet.row_values(i)
                    sample_name_value = str(row_data[sample_name_col_idx]).strip()
                    
                    # 删除含有'SB'关键词的行
                    if 'SB' not in sample_name_value:
                        table_b_rows.append(row_data)
                
                # 获取curve_points值（标准曲线点数）
                curve_points = config.curve_points
                n = curve_points + 1
                
                # 准备要插入的行
                insert_rows = []
                
                # 第一部分：3行SB
                for _ in range(3):
                    sb_row = [''] * gzkm_ncols
                    sb_row[sample_name_col_idx] = 'SB'
                    insert_rows.append(sb_row)
                
                # 第二部分：n行STD0-STD{n-1}
                for i in range(n):
                    std_row = [''] * gzkm_ncols
                    std_row[sample_name_col_idx] = f'STD{i}'
                    insert_rows.append(std_row)
                
                # 第三部分：2行SB
                for _ in range(2):
                    sb_row = [''] * gzkm_ncols
                    sb_row[sample_name_col_idx] = 'SB'
                    insert_rows.append(sb_row)
                
                # 将插入的行添加到表格B的开头
                final_rows = insert_rows + table_b_rows
                
                # 读取该项目的对应关系表（mapping_file）
                mapping_dict = {}
                if config.mapping_file:
                    try:
                        mapping_path = config.mapping_file.path
                        if os.path.exists(mapping_path):
                            # 读取Excel映射文件
                            mapping_df = pd.read_excel(mapping_path, sheet_name='上机列表')
                            
                            # 映射文件格式为：第一列是样品名称关键词，其他列是对应的值
                            # 根据实际映射文件格式调整
                            for _, row in mapping_df.iterrows():
                                key = str(row.iloc[0]).strip()  # 第一列作为匹配键
                                mapping_dict[key] = {
                                    gzkm_header[i]: str(row.iloc[i]) if i < len(row) else ''
                                    for i in range(1, len(gzkm_header))
                                }
                    except Exception as e:
                        print(f"读取映射文件失败: {e}")
                
                # 依据对应关系表，匹配补齐空白列的内容
                for row in final_rows:
                    sample_name_value = str(row[sample_name_col_idx]).strip()
                    
                    # 遍历映射字典，寻找匹配项
                    for key, mapping_values in mapping_dict.items():
                        if key in sample_name_value or sample_name_value in key:
                            # 匹配成功，填充其他列的空白内容
                            for col_idx, col_name in enumerate(gzkm_header):
                                if col_idx != sample_name_col_idx:
                                    # 如果当前单元格为空，则填充映射值
                                    if not row[col_idx] or str(row[col_idx]).strip() == '':
                                        if col_name in mapping_values:
                                            row[col_idx] = mapping_values[col_name]
                            break
                
                # 构建worklist_records数据结构
                for row in final_rows:
                    record = {}
                    for col_idx, col_name in enumerate(gzkm_header):
                        record[col_name] = str(row[col_idx]) if col_idx < len(row) else ''
                    worklist_records.append(record)


    # ========== 6. 构建 plates 数据结构（用于模板渲染）==========
    worksheet_table = [[worksheet_grid[r][c] for c in range(12)] for r in range(8)]
    
    plates = [{
        "header": {
            "plate_no": "X1",
            "instrument_num": instrument_num,
            "systerm_num": systerm_num,
            "injection_plate": "",  # 全血工作站无上机盘号
        },
        "worksheet_table": worksheet_table,
        "worksheet_table_2": worksheet_table_2,   # ★ 新增
        "error_rows": error_rows,
        "txt_headers": txt_headers,  # 添加上机列表表头
        "worklist_records": worklist_records,  # 添加上机列表数据
    }]
    
    # ========== 7. 将数据保存到 session（供导出函数使用）==========
    payload = {
        "project_name": project_name,
        "project_name_full": project_name_full,
        "instrument_num": instrument_num,
        "systerm_num": systerm_num,
        "platform": platform,
        "today_str": today_str,
        "plates": [{
            "plate_no": "X1",
            "worksheet_table": worksheet_table,
            "worksheet_table_2": worksheet_table_2,   # ★ 新增
            "error_rows": error_rows,
            "txt_headers": txt_headers,  # 添加上机列表表头
            "worklist_records": worklist_records,  # 添加上机列表数据
        }],
    }
    
    request.session['wholeblood_payload'] = payload
    request.session.modified = True
    
    # ========== 8. 返回中间展示页面 ==========
    return render(request, "dashboard/ProcessResult_WholeBloodWorkstation.html", {
        "project_name": project_name,
        "project_name_full": project_name_full,
        "instrument_num": instrument_num,
        "systerm_num": systerm_num,
        "today_str": today_str[:4] + "-" + today_str[4:6] + "-" + today_str[6:8],
        "platform": platform,
        "plates": plates,
    })


# ========== 第二步：用户点击"导出本板"时，生成文件并保存到服务器 ==========
@require_POST
def export_wholeblood_files(request):
    """
    导出全血工作站工作清单和报错信息表（第二步）
    """
    # 从 session 中读取之前保存的数据
    payload = request.session.get('wholeblood_payload')
    if not payload:
        return JsonResponse({"ok": False, "message": "数据已过期，请重新上传文件"}, status=400)
    
    project_name = payload['project_name']
    project_name_full = payload['project_name_full']
    instrument_num = payload['instrument_num']
    systerm_num = payload['systerm_num']
    platform = payload['platform']
    today_str = payload['today_str']
    plates = payload['plates']
    
    # 获取要导出的板号（支持多板，但全血工作站目前只有一块板）
    plate_index = int(request.GET.get('plate', 0))
    if plate_index >= len(plates):
        return JsonResponse({"ok": False, "message": "板号不存在"}, status=400)
    
    plate_data = plates[plate_index]
    plate_no = plate_data['plate_no']
    worksheet_table = plate_data['worksheet_table']
    error_rows = plate_data['error_rows']
    
    # ========== 1. 生成时间戳和文件名 ==========
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    worksheet_filename_stem = f"{plate_no}_WorkSheet_{instrument_num}_{systerm_num}_{project_name}_{timestamp}_{plate_no}_GZ"
    
    # ========== 2. 创建保存目录 ==========
    date_folder = today_str[:4] + "-" + today_str[4:6] + "-" + today_str[6:8]
    save_dir = os.path.join(
        settings.DOWNLOAD_ROOT,
        platform,
        date_folder,
        project_name
    )
    os.makedirs(save_dir, exist_ok=True)
    
    # ========== 3. 生成 PDF 工作清单 ==========
    pdf_payload = {
        "project_name": project_name,
        "project_name_full": project_name_full,
        "instrument_num": instrument_num,
        "systerm_num": systerm_num,
        "plate_no": plate_no,
        "worksheet_table": worksheet_table,
        "error_rows": error_rows,
        "platform": platform,
    }
    
    # 渲染 HTML 模板
    worksheet_html = render_to_string("dashboard/export_pdf.html", pdf_payload)
    
    # 配置字体和样式
    font_config = FontConfiguration()
    pdf_css = CSS(string="""
        @page { size: A4; margin: 10mm; }
        body { font-family: "Noto Sans CJK SC", "SimSun", sans-serif; font-size: 10px; }
        .highlight { background-color: #ffcccc; }
    """, font_config=font_config)
    
    # 生成 PDF 文件
    worksheet_pdf_filename = f"{worksheet_filename_stem}.pdf"
    worksheet_pdf_path = os.path.join(save_dir, worksheet_pdf_filename)
    
    HTML(string=worksheet_html).write_pdf(
        worksheet_pdf_path,
        stylesheets=[pdf_css],
        font_config=font_config
    )
    
    # ========== 4. 保存 payload.json（用于后续重新生成）==========
    payload_filename = f"{worksheet_filename_stem}.payload.json"
    payload_path = os.path.join(save_dir, payload_filename)
    
    with open(payload_path, "w", encoding="utf-8") as f:
        json.dump(pdf_payload, f, ensure_ascii=False, indent=2)
    
    # ========== 5. 生成 Excel 报错信息表（如果有报错）==========
    if error_rows:
        error_wb = xlwt.Workbook()
        error_sheet = error_wb.add_sheet("报错信息")
        
        # 表头
        headers = ["实验号", "条码", "板号", "孔位", "警告信息"]
        for col, header in enumerate(headers):
            error_sheet.write(0, col, header)
        
        # 数据行
        for row_idx, err in enumerate(error_rows, start=1):
            error_sheet.write(row_idx, 0, err["sample_name"])
            error_sheet.write(row_idx, 1, err["origin_barcode"])
            error_sheet.write(row_idx, 2, err["plate_no"])
            error_sheet.write(row_idx, 3, err["well_str"])
            error_sheet.write(row_idx, 4, err["warn_level"])
            error_sheet.write(row_idx, 5, err["warn_info"])
        
        error_filename = f"{plate_no}_ErrorList_{instrument_num}_{systerm_num}_{project_name}_{timestamp}_{plate_no}_GZ.xls"
        error_path = os.path.join(save_dir, error_filename)
        error_wb.save(error_path)
    

    # ========== 6. 生成上机列表 Excel 文件（如果有数据）==========
    txt_headers = plate_data.get('txt_headers', [])
    worklist_records = plate_data.get('worklist_records', [])
    
    if worklist_records and txt_headers:
        worklist_wb = xlwt.Workbook()
        worklist_sheet = worklist_wb.add_sheet("上机列表")
        
        # 写入表头
        for col, header in enumerate(txt_headers):
            worklist_sheet.write(0, col, header)
        
        # 写入数据行
        for row_idx, record in enumerate(worklist_records, start=1):
            for col_idx, col_name in enumerate(txt_headers):
                value = record.get(col_name, '')
                worklist_sheet.write(row_idx, col_idx, value)
        
        # 保存上机列表文件
        worklist_filename = f"{plate_no}_InjectionList_{instrument_num}_{systerm_num}_{project_name}_{timestamp}_{plate_no}_GZ.xls"
        worklist_path = os.path.join(save_dir, worklist_filename)
        worklist_wb.save(worklist_path)

    # ========== 6. 返回成功响应 ==========
    return JsonResponse({"ok": True, "message": "导出成功"})
